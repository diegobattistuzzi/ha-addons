import csv
import io
import re
import zipfile
from datetime import date as date_cls, datetime
from typing import Any, Dict, List, Optional, Tuple

from dateutil.parser import parse as parse_date


def parse_amount(value: Any) -> Optional[float]:
    if value is None or str(value).strip() == '':
        return None
    text = str(value).strip().replace(' ', '').replace('€', '').replace("'", '')
    if not text or text in ('-', '+'):
        return None
    if text.endswith('-'):
        text = '-' + text[:-1]
    # Italiano con separatore migliaia: 1.234.567,89 oppure 1.234
    if re.match(r'^-?\d{1,3}(\.\d{3})+(,\d+)?$', text):
        text = text.replace('.', '').replace(',', '.')
    # Inglese con separatore migliaia: 1,234.56 oppure 1234.56 oppure 1,234
    elif re.match(r'^-?\d{1,3}(,\d{3})*(\.\d+)?$', text):
        text = text.replace(',', '')
    else:
        text = text.replace(',', '.')
    try:
        return float(text)
    except ValueError:
        return None


def parse_rows_from_csv(data: bytes, sign_mode: str = 'auto') -> List[Dict[str, Any]]:
    text = data.decode('utf-8-sig', errors='replace')
    first_line = text.split('\n', 1)[0]
    delimiter = ';' if first_line.count(';') > first_line.count(',') else ','
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    return parse_tabular_rows(rows, sign_mode)


def parse_rows_from_xlsx(data: bytes, sign_mode: str = 'auto') -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return parse_tabular_rows(rows, sign_mode)


_XLSX_NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'


def _read_xlsx_raw_rows(data: bytes) -> List[List[Optional[str]]]:
    """Legge le celle del primo foglio di un .xlsx leggendo direttamente l'XML
    interno, senza passare da openpyxl: openpyxl.load_workbook rifiuta con un
    TypeError alcuni file .xlsx generati da tool terzi il cui stylesheet
    (styles.xml) e' tecnicamente non valido (bug reale trovato su un vero
    export buoni pasto - "Fill() takes no arguments"). Qui interessa solo il
    contenuto delle celle, non la formattazione, quindi gli stili non vengono
    proprio toccati."""
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(io.BytesIO(data)) as z:
        sheet_names = sorted(n for n in z.namelist() if re.match(r'xl/worksheets/sheet\d+\.xml$', n))
        if not sheet_names:
            raise ValueError('Nessun foglio trovato nel file xlsx')
        shared_strings: List[str] = []
        if 'xl/sharedStrings.xml' in z.namelist():
            sst_root = ET.fromstring(z.read('xl/sharedStrings.xml'))
            for si in sst_root.findall(f'{_XLSX_NS}si'):
                shared_strings.append(''.join(t.text or '' for t in si.iter(f'{_XLSX_NS}t')))
        root = ET.fromstring(z.read(sheet_names[0]))
        rows: List[List[Optional[str]]] = []
        for row_el in root.iter(f'{_XLSX_NS}row'):
            cells: List[Optional[str]] = []
            for c in row_el.iter(f'{_XLSX_NS}c'):
                cell_type = c.get('t')
                if cell_type == 's':
                    v_el = c.find(f'{_XLSX_NS}v')
                    idx = int(v_el.text) if v_el is not None and v_el.text else None
                    cells.append(shared_strings[idx] if idx is not None and idx < len(shared_strings) else None)
                elif cell_type == 'inlineStr':
                    is_el = c.find(f'{_XLSX_NS}is')
                    t_el = is_el.find(f'{_XLSX_NS}t') if is_el is not None else None
                    cells.append(t_el.text if t_el is not None else None)
                else:
                    v_el = c.find(f'{_XLSX_NS}v')
                    cells.append(v_el.text if v_el is not None else None)
            rows.append(cells)
        return rows


_MEAL_VOUCHER_HEADER = {'data e ora', 'tipo movimento', 'supporto', 'n. e importo buoni', 'dettaglio'}


def looks_like_meal_voucher_export(data: bytes) -> bool:
    """Vero per l'export 'Elenco movimenti' dei buoni pasto (Edenred/Cloud e
    simili): un .xlsx con un tracciato particolare a blocchi di 4 righe per
    movimento (intestazione ripetuta + riga dati + sotto-intestazione + riga
    dettaglio) che il parser xlsx generico (una sola intestazione, poi una
    riga per movimento) non sa interpretare - vedi parse_rows_from_meal_voucher."""
    try:
        rows = _read_xlsx_raw_rows(data)
    except Exception:
        return False
    if not rows:
        return False
    header = {str(cell).strip().lower() for cell in rows[0] if cell}
    return _MEAL_VOUCHER_HEADER.issubset(header)


_MEAL_VOUCHER_COUNT_PRICE_RE = re.compile(r'(\d+)\s*da\s*[^\d]*([\d.,]+)')
_MEAL_VOUCHER_MERCHANT_RE = re.compile(r'presso\s+(.+)$', re.IGNORECASE)


def parse_rows_from_meal_voucher(data: bytes) -> List[Dict[str, Any]]:
    """Converte l'export 'Elenco movimenti' dei buoni pasto nello stesso
    formato {'date','amount','description'} degli altri importatori. Ogni
    movimento occupa un blocco di 4 righe (intestazione ripetuta, riga dati,
    sotto-intestazione, riga dettaglio): qui interessano solo intestazione e
    riga dati, le altre due (metadati del carnet/esercente) sono ignorate.
    Gli 'Ordine Cloud' (ricarica) diventano accrediti, gli 'Utilizzo' spese.
    Piu' righe 'Utilizzo' con la stessa data/ora e lo stesso esercente (una
    spesa pagata impilando piu' buoni, es. la spesa al supermercato) vengono
    sommate in un'unica transazione, altrimenti una spesa reale finirebbe
    spezzata in N righe identiche da poche euro l'una."""
    rows = _read_xlsx_raw_rows(data)
    entries: List[Dict[str, Any]] = []

    i = 0
    while i < len(rows):
        header_cells = {str(cell).strip().lower() for cell in rows[i] if cell}
        if not _MEAL_VOUCHER_HEADER.issubset(header_cells):
            i += 1
            continue
        if i + 1 >= len(rows):
            break
        data_row = rows[i + 1]
        i += 4  # blocco intero: riga dati + sotto-intestazione + sotto-dati

        when_raw, tipo, _supporto, count_price, dettaglio = (list(data_row) + [None] * 5)[:5]
        if not when_raw or not tipo or not count_price:
            continue
        m = _MEAL_VOUCHER_COUNT_PRICE_RE.search(str(count_price))
        if not m:
            continue
        count = int(m.group(1))
        unit_price = parse_amount(m.group(2))
        if unit_price is None:
            continue
        try:
            when = datetime.strptime(str(when_raw).strip(), '%d/%m/%Y %H:%M:%S')
        except ValueError:
            continue

        tipo_lower = tipo.strip().lower()
        dettaglio = (dettaglio or '').strip()
        if 'ordine' in tipo_lower:
            entries.append({
                'when': when, 'signed_total': count * unit_price,
                'count': count, 'unit_price': unit_price, 'label': 'Accredito buoni pasto',
            })
        elif 'utilizzo' in tipo_lower:
            merchant_match = _MEAL_VOUCHER_MERCHANT_RE.search(dettaglio)
            merchant = merchant_match.group(1).strip() if merchant_match else (dettaglio or 'Buoni pasto')
            entries.append({
                'when': when, 'signed_total': -count * unit_price,
                'count': count, 'unit_price': unit_price, 'label': merchant,
            })
        # altri 'Tipo movimento' non ancora visti in un export reale (es.
        # eventuali storni) vengono ignorati piuttosto che interpretati alla cieca.

    grouped: Dict[Tuple[datetime, str], Dict[str, Any]] = {}
    for e in entries:
        key = (e['when'], e['label'])
        g = grouped.setdefault(key, {'label': e['label'], 'when': e['when'], 'total': 0.0, 'count': 0, 'unit_price': e['unit_price']})
        g['total'] += e['signed_total']
        g['count'] += e['count']

    result = []
    for g in grouped.values():
        plural = 'o' if g['count'] == 1 else 'i'
        description = f"{g['label']} ({g['count']} buon{plural} da {g['unit_price']:.2f}€)"
        result.append({
            'date': g['when'].date().isoformat(),
            'amount': round(g['total'], 2),
            'description': description,
        })
    return result


def looks_like_cbi(data: bytes) -> bool:
    """Vero se il file e' un tracciato CBI a record fissi (righe di 120
    caratteri, header 'RH' come prima riga non vuota) invece di un CSV: alcune
    banche esportano i movimenti conto corrente in questo formato proprietario
    con estensione .txt, che altrimenti finirebbe (fallendo) nel parser CSV."""
    text = data[:200].decode('latin-1', errors='replace')
    first_line = text.split('\r\n', 1)[0].split('\n', 1)[0].split('\r', 1)[0]
    return bool(re.match(r'^\s?RH\d', first_line))


def parse_rows_from_cbi(data: bytes) -> Tuple[Optional[Dict[str, Any]], List[Dict[str, Any]]]:
    """Tracciato CBI a record fissi (120 caratteri) usato da diverse banche
    italiane per l'export dei movimenti conto corrente: un record '61' apre
    ogni giornata (saldo iniziale + IBAN del conto), seguito da zero o piu'
    coppie di record '62' (un movimento ciascuno, con data/importo/segno/
    causale) e '63' (righe di dettaglio testuale legate all'ultimo '62': nome/
    indirizzo ordinante o causale estesa). Chiudono la giornata i record '64'
    (saldo finale) e '65' (valute in maturazione nei giorni successivi),
    ignorati perche' non sono movimenti. Gli offset sono dedotti da un unico
    export reale (banca Intesa Sanpaolo-CBI): se un altro export usa layout
    leggermente diverso questa funzione potrebbe interpretare male i campi
    invece di segnalare un errore, dato che il formato non e' documentato
    pubblicamente in modo univoco."""
    text = data.decode('latin-1', errors='replace')
    lines = [line for line in re.split(r'\r\n|\r|\n', text) if line.strip()]

    detected_account: Optional[Dict[str, Any]] = None
    rows: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    current_key = None

    for line in lines:
        if len(line) < 45:
            continue
        rectype = line[1:3]

        if rectype == '61':
            if detected_account is None:
                iban_body = line[51:74].strip()
                country = line[99:103].strip()
                if iban_body and country:
                    detected_account = {'iban': f'{country}{iban_body}'}
            continue

        if rectype == '62':
            current = None
            current_key = None
            date_op = line[13:19]
            sign = line[25]
            try:
                date_text = datetime.strptime(date_op, '%d%m%y').date().isoformat()
            except ValueError:
                continue
            amount = parse_amount(line[26:41])
            if amount is None:
                continue
            amount = -abs(amount) if sign == 'D' else abs(amount)
            description = line[86:120].strip() or 'Movimento CBI'
            current = {'date': date_text, 'amount': amount, 'description': description}
            current_key = (line[3:10], line[10:13])
            rows.append(current)
            continue

        if rectype == '63':
            if current is None or current_key != (line[3:10], line[10:13]):
                continue
            extra = line[13:].rstrip()
            # 'YYY'/'YY2' (o simili) sono righe strutturate nome/indirizzo
            # ordinante duplicate rispetto alla causale, prive di separatore
            # dopo il codice ("YYY13032023...", "YY2PIAZZA...") - a differenza
            # di testo libero che inizia per coincidenza con 2 lettere
            # maiuscole ma seguite da un separatore reale (es. "ASS: ...").
            if re.match(r'^YY[Y\d]', extra):
                continue
            extra = extra.strip()
            extra = re.sub(r'^\*', '', extra)
            extra = re.sub(r'^\d{4}-\d{2}-\d{2}\*', '', extra)
            extra = re.sub(r'\s+', ' ', extra).strip()
            if extra and extra.lower() not in current['description'].lower():
                current['description'] = f"{current['description']} - {extra}".strip(' -')
            continue

        # '64' (saldo di chiusura) e '65' (valute future): non sono movimenti.

    return detected_account, rows


# Quante righe di preambolo tollerare prima di rinunciare a trovare l'header
# vero: alcuni export bancari (es. Intesa/Moneymap) anteponono un blocco di
# metadati (numero conto, intestatario, periodo, saldo finale, note legali,
# righe vuote, un titolo tipo "Risultati Ricerca") PRIMA della vera riga di
# intestazione colonne - bug reale trovato su un export reale, dove
# assumere che rows[0] fosse sempre l'header faceva fallire il riconoscimento
# della colonna data/importo (nessuna delle due esiste nella prima riga, che
# e' solo "Conto Corrente: 123456") con un errore 500 non gestito.
_HEADER_SCAN_ROWS = 30


def _detect_columns(normalized_headers: List[str]):
    """Individua gli indici delle colonne note a partire dagli header gia'
    normalizzati (minuscolo). 'uscit'/'entrat' (Uscite/Entrate) sono sinonimi
    reali di addebito/accredito su almeno un export bancario italiano visto,
    dove le due colonne separate si chiamano cosi' e non 'debito'/'credito'.
    description_idx esclude gli indici gia' assegnati a date/amount/debit/
    credit: senza questa esclusione, un header come "Data_Operazione"
    veniva scambiato per la colonna descrizione (matcha 'operazione', uno
    dei token di descrizione) PRIMA di arrivare alla vera colonna
    "Descrizione" - bug reale trovato sullo stesso export."""
    date_idx = next((i for i, h in enumerate(normalized_headers) if any(t in h for t in ['date', 'data'])), None)
    # Un rendiconto carta puo' avere DUE colonne "importo": quella nella
    # valuta originale dell'acquisto (vuota per le spese gia' in euro,
    # valorizzata solo per quelle in valuta estera) e quella vera nella
    # valuta del conto - bug reale trovato su un export reale, dove
    # "Importo originale in Divisa" veniva scelta perche' matcha 'import' per
    # prima, lasciando None (quindi riga scartata) su tutte le spese in euro
    # e facendo sopravvivere solo quelle in valuta estera. Scartiamo prima le
    # colonne "originale/divisa" e usiamo quelle solo se non c'e' alternativa.
    amount_candidates = [i for i, h in enumerate(normalized_headers) if any(t in h for t in ['amount', 'import', 'totale', 'valore'])]
    amount_idx = next((i for i in amount_candidates if not any(t in normalized_headers[i] for t in ['divisa', 'original'])), None)
    if amount_idx is None:
        amount_idx = amount_candidates[0] if amount_candidates else None
    debit_idx = next((i for i, h in enumerate(normalized_headers) if any(t in h for t in ['debit', 'addebito', 'uscit'])), None)
    credit_idx = next((i for i, h in enumerate(normalized_headers) if any(t in h for t in ['credit', 'accredito', 'entrat'])), None)
    taken = {i for i in (date_idx, amount_idx, debit_idx, credit_idx) if i is not None}
    # Priorita' a una colonna "descrizione completa/estesa" quando esiste
    # accanto a una "descrizione" generica: alcuni estratti (visto su un
    # export reale) hanno entrambe, dove quella generica e' solo il tipo di
    # movimento ("VISA DEBIT", "Pagamento Visa Debit") e quella completa ha il
    # vero esercente/causale - usare la prima trovata darebbe sempre la meno
    # utile delle due, dato che di norma la generica compare per prima.
    description_idx = next(
        (i for i, h in enumerate(normalized_headers)
         if i not in taken and any(t in h for t in ['completa', 'estesa', 'extended', 'complete'])),
        None,
    )
    if description_idx is None:
        description_idx = next(
            (i for i, h in enumerate(normalized_headers)
             if i not in taken and any(t in h for t in ['desc', 'causale', 'merchant', 'note', 'azienda', 'operazione'])),
        None,
    )
    return date_idx, amount_idx, debit_idx, credit_idx, description_idx


def parse_tabular_rows(rows: List[List[Any]], sign_mode: str = 'auto') -> List[Dict[str, Any]]:
    rows = [row for row in rows if any(cell is not None and str(cell).strip() != '' for cell in row)]
    if not rows:
        return []

    # Cerca la vera riga di intestazione invece di assumere che sia la prima:
    # la prima riga che, una volta normalizzata, individua sia una colonna
    # data sia una colonna importo/debito/credito. Due controlli di
    # plausibilita' in piu', altrimenti una riga di note discorsive nel
    # preambolo (vedi _HEADER_SCAN_ROWS) puo' combaciare per puro caso di
    # sottostringa - bug reale trovato su un export reale, dove una nota
    # legale che cita "carte di credito" e "data valuta" nello stesso
    # paragrafo veniva scambiata per intestazione:
    # 1) gli indici individuati devono essere tutti DISTINTI - un vero header
    #    non puo' avere la stessa cella che e' "la colonna data" E "la
    #    colonna credito" insieme, mentre una frase in un'unica cella si'
    #    (l'intera frase finisce nella stessa colonna 0);
    # 2) la riga deve avere piu' di una cella non vuota - un vero header
    #    elenca piu' nomi di colonna, una nota e' un'unica frase lunga in
    #    una sola cella con il resto della riga vuoto.
    header_row_idx = None
    date_idx = amount_idx = debit_idx = credit_idx = description_idx = None
    for i, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        non_empty_cells = sum(1 for cell in row if cell is not None and str(cell).strip() != '')
        if non_empty_cells < 2:
            continue
        normalized = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        d_idx, a_idx, deb_idx, cred_idx, desc_idx = _detect_columns(normalized)
        found = [idx for idx in (d_idx, a_idx, deb_idx, cred_idx) if idx is not None]
        if d_idx is not None and (a_idx is not None or deb_idx is not None or cred_idx is not None) and len(found) == len(set(found)):
            header_row_idx = i
            date_idx, amount_idx, debit_idx, credit_idx, description_idx = d_idx, a_idx, deb_idx, cred_idx, desc_idx
            break

    if header_row_idx is None:
        raise ValueError(
            "Intestazione non riconosciuta: non trovo una colonna data insieme a una colonna importo/entrate/uscite "
            f'nelle prime {_HEADER_SCAN_ROWS} righe del file.'
        )

    # Un rendiconto di carta di credito ha spesso UNA sola colonna importo,
    # sempre positiva (ogni riga e' un acquisto), senza alcun segno esplicito
    # ne' colonne addebiti/accrediti separate - bug reale trovato su un export
    # reale ("Movimenti Carta ****", "Carta di Credito" nel preambolo, colonna
    # unica "Importo (€)" sempre positiva anche per spese vere). Lo stesso
    # principio gia' usato per i PDF (vedi pdf_import.is_credit_card_statement)
    # qui si applica solo quando il segno arriva dalla colonna amount_idx
    # (non da un debit/credit gia' esplicito): solo se il valore e' positivo,
    # per non invertire un eventuale storno/rimborso gia' scritto negativo.
    # sign_mode arriva dal conto selezionato per l'import (vedi accounts.
    # amount_sign_mode in migrate.py): quando l'utente lo ha fissato
    # esplicitamente per un istituto la cui euristica sul preambolo sbaglia
    # (es. American Express, che gia' usa spese negative/accrediti positivi
    # invece dell'unica-colonna-sempre-positiva assunta qui sotto), quel
    # valore vince e il testo del preambolo non viene nemmeno guardato.
    if sign_mode == 'flip':
        is_credit_card_statement = True
    elif sign_mode == 'signed':
        is_credit_card_statement = False
    else:
        preamble_text = ' '.join(
            str(cell).lower() for row in rows[:header_row_idx] for cell in row if cell is not None
        )
        is_credit_card_statement = bool(re.search(
            r'carta\s+di\s+credito|carta\s+revolving|movimenti\s+carta|limite\s+di\s+utilizzo', preamble_text,
        ))

    # Colonne diverse possono contenere una data valida per righe diverse
    # (es. "Data_Operazione" vale "-" per i movimenti carta non ancora
    # contabilizzati, dove solo "Data_Valuta" e' una data vera): proviamo
    # TUTTE le colonne che sembrano una data, non solo quella scelta come
    # riferimento principale, invece di scartare silenziosamente la riga.
    date_candidate_idxs = [date_idx] + [
        i for i, h in enumerate(
            [str(c).strip().lower() if c is not None else '' for c in rows[header_row_idx]]
        )
        if i != date_idx and any(t in h for t in ['date', 'data'])
    ]
    # Colonna "Data Valuta" separata da quella scelta come data principale
    # (operazione): se presente, la teniamo per popolare value_date invece di
    # scartarla, cosi' il dedup puo' confrontarla oltre alla finestra
    # euristica di 3 giorni (vedi GET /api/transactions/duplicates).
    value_date_idx = next(
        (i for i, h in enumerate(
            [str(c).strip().lower() if c is not None else '' for c in rows[header_row_idx]]
        )
        if i != date_idx and 'valuta' in h),
        None,
    )

    def _extract_date(row: List[Any], idxs: List[int]) -> Optional[str]:
        for idx in idxs:
            if idx is None or idx >= len(row):
                continue
            value = row[idx]
            if value is None or str(value).strip() == '':
                continue
            try:
                if isinstance(value, date_cls):
                    # openpyxl restituisce gia' un datetime.date/datetime vero per
                    # le celle formattate come data in Excel: usarlo direttamente,
                    # SENZA passare da str(value) + dateutil, e' l'unico modo
                    # sicuro di evitare un bug reale trovato importando un vero
                    # estratto - str(datetime(2026,7,1)) produce "2026-07-01
                    # 00:00:00", e dateutil con dayfirst=True (necessario per le
                    # date scritte a mano in formato italiano nei CSV) scambia
                    # comunque giorno e mese quando il giorno reale e' <=12 anche
                    # se l'anno a 4 cifre rende la stringa gia' inequivocabile
                    # (risultato: 1 luglio letto come 7 gennaio). Le date scritte
                    # come testo (CSV, o celle Excel non formattate come data)
                    # restano sul percorso dateutil sotto, dove dayfirst serve
                    # davvero.
                    return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
                return parse_date(str(value), dayfirst=True).date().isoformat()
            except Exception:
                continue
        return None

    result = []
    for row in rows[header_row_idx + 1:]:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        date_text = _extract_date(row, date_candidate_idxs)
        if date_text is None:
            continue
        value_date_text = _extract_date(row, [value_date_idx]) if value_date_idx is not None else None
        if value_date_text == date_text:
            value_date_text = None

        amount = None
        if amount_idx is not None and amount_idx < len(row):
            amount = parse_amount(row[amount_idx])
            if amount is not None and amount > 0 and is_credit_card_statement:
                amount = -amount
        if amount is None:
            debit = parse_amount(row[debit_idx]) if debit_idx is not None and debit_idx < len(row) else None
            credit = parse_amount(row[credit_idx]) if credit_idx is not None and credit_idx < len(row) else None
            if debit is not None and credit is None:
                amount = -abs(debit)
            elif credit is not None and debit is None:
                amount = abs(credit)
            elif debit is not None and credit is not None:
                amount = abs(credit) - abs(debit)
        if amount is None:
            continue

        description = ''
        if description_idx is not None and description_idx < len(row):
            description = str(row[description_idx] or '').strip()
        else:
            description = 'Importazione'

        result.append({
            'date': date_text,
            'value_date': value_date_text,
            'amount': amount,
            'description': description,
        })
    return result
