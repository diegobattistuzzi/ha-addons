import csv
import hashlib
import io
import json
import re
import uuid
import zipfile
from datetime import date as date_cls, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import httpx
from dateutil.parser import parse as parse_date
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.exceptions import HTTPException as StarletteHTTPException

from . import access, ai_client, backup, categorize, config, db, email_backfill, email_enrich, email_poller, pdf_import
from .migrate import run_migrations

BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / 'public'

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_methods=['*'],
    allow_headers=['*'],
)


@app.middleware('http')
async def enforce_public_gateway_auth(request: Request, call_next):
    """Chiude l'accesso anonimo alle API quando gira come add-on HA (unica
    situazione in cui puo' esistere una porta pubblica per l'uso mobile).

    L'Ingress di HA e' l'UNICO percorso fidato senza token: il Supervisor
    inietta l'header 'X-Ingress-Path' quando fa da proxy, un client esterno
    che raggiunge l'add-on da un'altra strada (es. la porta pubblica dietro
    nginx per la PWA mobile) non puo' impostarlo lui stesso SE nginx lo
    rimuove dagli header in ingresso (vedi README.md) - per questo motivo
    NON e' sufficiente controllare X-Remote-User-Id/X-Person-Id (quelli si
    possono spedire da chiunque): qui serve o l'Ingress genuino o un token
    mobile valido, altrimenti chiunque su internet leggerebbe conti e spese
    condivise senza autenticarsi. In sviluppo locale (fuori da HA, niente
    SUPERVISOR_TOKEN) il controllo resta disattivato.
    """
    if config.SUPERVISOR_TOKEN and request.url.path.startswith('/api'):
        if not request.headers.get('x-ingress-path'):
            if access.get_person_from_bearer(request) is None:
                return JSONResponse(status_code=401, content={'detail': 'Autenticazione richiesta'})
    return await call_next(request)


@app.on_event('startup')
def startup():
    try:
        run_migrations()
    except Exception as e:
        print(f'[startup] ERRORE migrazione: {e}', flush=True)
        raise
    email_poller.start_background_poller()


def row_to_dict(row: Optional[db.sqlite3.Row]) -> Optional[Dict[str, Any]]:
    if row is None:
        return None
    return {k: row[k] for k in row.keys()}


def fetchall(query: str, args: tuple = ()) -> List[Dict[str, Any]]:
    cursor = db.conn.execute(query, args)
    return [row_to_dict(row) for row in cursor.fetchall()]


def fetchone(query: str, args: tuple = ()) -> Optional[Dict[str, Any]]:
    cursor = db.conn.execute(query, args)
    return row_to_dict(cursor.fetchone())


def execute(query: str, args: tuple = ()) -> int:
    cursor = db.conn.execute(query, args)
    db.conn.commit()
    return cursor.rowcount


def parse_amount(value: Any) -> Optional[float]:
    if value is None or str(value).strip() == '':
        return None
    text = str(value).strip().replace(' ', '').replace('€', '').replace("'", '')
    if not text or text in ('-', '+'):
        return None
    if text.endswith('-'):
        text = '-' + text[:-1]
    # Italiano con separatore migliaia: 1.234.567,89 oppure 1.234
    if re.match(r'^-?\d{1,3}(\.\d{3})+(,\d+)?$', text):
        text = text.replace('.', '').replace(',', '.')
    # Inglese con separatore migliaia: 1,234.56 oppure 1234.56 oppure 1,234
    elif re.match(r'^-?\d{1,3}(,\d{3})*(\.\d+)?$', text):
        text = text.replace(',', '')
    else:
        text = text.replace(',', '.')
    try:
        return float(text)
    except ValueError:
        return None


def parse_rows_from_csv(data: bytes) -> List[Dict[str, Any]]:
    text = data.decode('utf-8-sig', errors='replace')
    first_line = text.split('\n', 1)[0]
    delimiter = ';' if first_line.count(';') > first_line.count(',') else ','
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    return parse_tabular_rows(rows)


def parse_rows_from_xlsx(data: bytes) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return parse_tabular_rows(rows)


_XLSX_NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'


def _read_xlsx_raw_rows(data: bytes) -> List[List[Optional[str]]]:
    """Legge le celle del primo foglio di un .xlsx leggendo direttamente l'XML
    interno, senza passare da openpyxl: openpyxl.load_workbook rifiuta con un
    TypeError alcuni file .xlsx generati da tool terzi il cui stylesheet
    (styles.xml) e' tecnicamente non valido (bug reale trovato su un vero
    export buoni pasto - "Fill() takes no arguments"). Qui interessa solo il
    contenuto delle celle, non la formattazione, quindi gli stili non vengono
    proprio toccati."""
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(io.BytesIO(data)) as z:
        sheet_names = sorted(n for n in z.namelist() if re.match(r'xl/worksheets/sheet\d+\.xml$', n))
        if not sheet_names:
            raise ValueError('Nessun foglio trovato nel file xlsx')
        shared_strings: List[str] = []
        if 'xl/sharedStrings.xml' in z.namelist():
            sst_root = ET.fromstring(z.read('xl/sharedStrings.xml'))
            for si in sst_root.findall(f'{_XLSX_NS}si'):
                shared_strings.append(''.join(t.text or '' for t in si.iter(f'{_XLSX_NS}t')))
        root = ET.fromstring(z.read(sheet_names[0]))
        rows: List[List[Optional[str]]] = []
        for row_el in root.iter(f'{_XLSX_NS}row'):
            cells: List[Optional[str]] = []
            for c in row_el.iter(f'{_XLSX_NS}c'):
                cell_type = c.get('t')
                if cell_type == 's':
                    v_el = c.find(f'{_XLSX_NS}v')
                    idx = int(v_el.text) if v_el is not None and v_el.text else None
                    cells.append(shared_strings[idx] if idx is not None and idx < len(shared_strings) else None)
                elif cell_type == 'inlineStr':
                    is_el = c.find(f'{_XLSX_NS}is')
                    t_el = is_el.find(f'{_XLSX_NS}t') if is_el is not None else None
                    cells.append(t_el.text if t_el is not None else None)
                else:
                    v_el = c.find(f'{_XLSX_NS}v')
                    cells.append(v_el.text if v_el is not None else None)
            rows.append(cells)
        return rows


_MEAL_VOUCHER_HEADER = {'data e ora', 'tipo movimento', 'supporto', 'n. e importo buoni', 'dettaglio'}


def looks_like_meal_voucher_export(data: bytes) -> bool:
    """Vero per l'export 'Elenco movimenti' dei buoni pasto (Edenred/Cloud e
    simili): un .xlsx con un tracciato particolare a blocchi di 4 righe per
    movimento (intestazione ripetuta + riga dati + sotto-intestazione + riga
    dettaglio) che il parser xlsx generico (una sola intestazione, poi una
    riga per movimento) non sa interpretare - vedi parse_rows_from_meal_voucher."""
    try:
        rows = _read_xlsx_raw_rows(data)
    except Exception:
        return False
    if not rows:
        return False
    header = {str(cell).strip().lower() for cell in rows[0] if cell}
    return _MEAL_VOUCHER_HEADER.issubset(header)


_MEAL_VOUCHER_COUNT_PRICE_RE = re.compile(r'(\d+)\s*da\s*[^\d]*([\d.,]+)')
_MEAL_VOUCHER_MERCHANT_RE = re.compile(r'presso\s+(.+)$', re.IGNORECASE)


def parse_rows_from_meal_voucher(data: bytes) -> List[Dict[str, Any]]:
    """Converte l'export 'Elenco movimenti' dei buoni pasto nello stesso
    formato {'date','amount','description'} degli altri importatori. Ogni
    movimento occupa un blocco di 4 righe (intestazione ripetuta, riga dati,
    sotto-intestazione, riga dettaglio): qui interessano solo intestazione e
    riga dati, le altre due (metadati del carnet/esercente) sono ignorate.
    Gli 'Ordine Cloud' (ricarica) diventano accrediti, gli 'Utilizzo' spese.
    Piu' righe 'Utilizzo' con la stessa data/ora e lo stesso esercente (una
    spesa pagata impilando piu' buoni, es. la spesa al supermercato) vengono
    sommate in un'unica transazione, altrimenti una spesa reale finirebbe
    spezzata in N righe identiche da poche euro l'una."""
    rows = _read_xlsx_raw_rows(data)
    entries: List[Dict[str, Any]] = []

    i = 0
    while i < len(rows):
        header_cells = {str(cell).strip().lower() for cell in rows[i] if cell}
        if not _MEAL_VOUCHER_HEADER.issubset(header_cells):
            i += 1
            continue
        if i + 1 >= len(rows):
            break
        data_row = rows[i + 1]
        i += 4  # blocco intero: riga dati + sotto-intestazione + sotto-dati

        when_raw, tipo, _supporto, count_price, dettaglio = (list(data_row) + [None] * 5)[:5]
        if not when_raw or not tipo or not count_price:
            continue
        m = _MEAL_VOUCHER_COUNT_PRICE_RE.search(str(count_price))
        if not m:
            continue
        count = int(m.group(1))
        unit_price = parse_amount(m.group(2))
        if unit_price is None:
            continue
        try:
            when = datetime.strptime(str(when_raw).strip(), '%d/%m/%Y %H:%M:%S')
        except ValueError:
            continue

        tipo_lower = tipo.strip().lower()
        dettaglio = (dettaglio or '').strip()
        if 'ordine' in tipo_lower:
            entries.append({
                'when': when, 'signed_total': count * unit_price,
                'count': count, 'unit_price': unit_price, 'label': 'Accredito buoni pasto',
            })
        elif 'utilizzo' in tipo_lower:
            merchant_match = _MEAL_VOUCHER_MERCHANT_RE.search(dettaglio)
            merchant = merchant_match.group(1).strip() if merchant_match else (dettaglio or 'Buoni pasto')
            entries.append({
                'when': when, 'signed_total': -count * unit_price,
                'count': count, 'unit_price': unit_price, 'label': merchant,
            })
        # altri 'Tipo movimento' non ancora visti in un export reale (es.
        # eventuali storni) vengono ignorati piuttosto che interpretati alla cieca.

    grouped: Dict[Tuple[datetime, str], Dict[str, Any]] = {}
    for e in entries:
        key = (e['when'], e['label'])
        g = grouped.setdefault(key, {'label': e['label'], 'when': e['when'], 'total': 0.0, 'count': 0, 'unit_price': e['unit_price']})
        g['total'] += e['signed_total']
        g['count'] += e['count']

    result = []
    for g in grouped.values():
        plural = 'o' if g['count'] == 1 else 'i'
        description = f"{g['label']} ({g['count']} buon{plural} da {g['unit_price']:.2f}€)"
        result.append({
            'date': g['when'].date().isoformat(),
            'amount': round(g['total'], 2),
            'description': description,
        })
    return result


def looks_like_cbi(data: bytes) -> bool:
    """Vero se il file e' un tracciato CBI a record fissi (righe di 120
    caratteri, header 'RH' come prima riga non vuota) invece di un CSV: alcune
    banche esportano i movimenti conto corrente in questo formato proprietario
    con estensione .txt, che altrimenti finirebbe (fallendo) nel parser CSV."""
    text = data[:200].decode('latin-1', errors='replace')
    first_line = text.split('\r\n', 1)[0].split('\n', 1)[0].split('\r', 1)[0]
    return bool(re.match(r'^\s?RH\d', first_line))


def parse_rows_from_cbi(data: bytes) -> Tuple[Optional[Dict[str, Any]], List[Dict[str, Any]]]:
    """Tracciato CBI a record fissi (120 caratteri) usato da diverse banche
    italiane per l'export dei movimenti conto corrente: un record '61' apre
    ogni giornata (saldo iniziale + IBAN del conto), seguito da zero o piu'
    coppie di record '62' (un movimento ciascuno, con data/importo/segno/
    causale) e '63' (righe di dettaglio testuale legate all'ultimo '62': nome/
    indirizzo ordinante o causale estesa). Chiudono la giornata i record '64'
    (saldo finale) e '65' (valute in maturazione nei giorni successivi),
    ignorati perche' non sono movimenti. Gli offset sono dedotti da un unico
    export reale (banca Intesa Sanpaolo-CBI): se un altro export usa layout
    leggermente diverso questa funzione potrebbe interpretare male i campi
    invece di segnalare un errore, dato che il formato non e' documentato
    pubblicamente in modo univoco."""
    text = data.decode('latin-1', errors='replace')
    lines = [line for line in re.split(r'\r\n|\r|\n', text) if line.strip()]

    detected_account: Optional[Dict[str, Any]] = None
    rows: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    current_key = None

    for line in lines:
        if len(line) < 45:
            continue
        rectype = line[1:3]

        if rectype == '61':
            if detected_account is None:
                iban_body = line[51:74].strip()
                country = line[99:103].strip()
                if iban_body and country:
                    detected_account = {'iban': f'{country}{iban_body}'}
            continue

        if rectype == '62':
            current = None
            current_key = None
            date_op = line[13:19]
            sign = line[25]
            try:
                date_text = datetime.strptime(date_op, '%d%m%y').date().isoformat()
            except ValueError:
                continue
            amount = parse_amount(line[26:41])
            if amount is None:
                continue
            amount = -abs(amount) if sign == 'D' else abs(amount)
            description = line[86:120].strip() or 'Movimento CBI'
            current = {'date': date_text, 'amount': amount, 'description': description}
            current_key = (line[3:10], line[10:13])
            rows.append(current)
            continue

        if rectype == '63':
            if current is None or current_key != (line[3:10], line[10:13]):
                continue
            extra = line[13:].rstrip()
            # 'YYY'/'YY2' (o simili) sono righe strutturate nome/indirizzo
            # ordinante duplicate rispetto alla causale, prive di separatore
            # dopo il codice ("YYY13032023...", "YY2PIAZZA...") - a differenza
            # di testo libero che inizia per coincidenza con 2 lettere
            # maiuscole ma seguite da un separatore reale (es. "ASS: ...").
            if re.match(r'^YY[Y\d]', extra):
                continue
            extra = extra.strip()
            extra = re.sub(r'^\*', '', extra)
            extra = re.sub(r'^\d{4}-\d{2}-\d{2}\*', '', extra)
            extra = re.sub(r'\s+', ' ', extra).strip()
            if extra and extra.lower() not in current['description'].lower():
                current['description'] = f"{current['description']} - {extra}".strip(' -')
            continue

        # '64' (saldo di chiusura) e '65' (valute future): non sono movimenti.

    return detected_account, rows


# Quante righe di preambolo tollerare prima di rinunciare a trovare l'header
# vero: alcuni export bancari (es. Intesa/Moneymap) anteponono un blocco di
# metadati (numero conto, intestatario, periodo, saldo finale, note legali,
# righe vuote, un titolo tipo "Risultati Ricerca") PRIMA della vera riga di
# intestazione colonne - bug reale trovato su un export reale, dove
# assumere che rows[0] fosse sempre l'header faceva fallire il riconoscimento
# della colonna data/importo (nessuna delle due esiste nella prima riga, che
# e' solo "Conto Corrente: 123456") con un errore 500 non gestito.
_HEADER_SCAN_ROWS = 30


def _detect_columns(normalized_headers: List[str]):
    """Individua gli indici delle colonne note a partire dagli header gia'
    normalizzati (minuscolo). 'uscit'/'entrat' (Uscite/Entrate) sono sinonimi
    reali di addebito/accredito su almeno un export bancario italiano visto,
    dove le due colonne separate si chiamano cosi' e non 'debito'/'credito'.
    description_idx esclude gli indici gia' assegnati a date/amount/debit/
    credit: senza questa esclusione, un header come "Data_Operazione"
    veniva scambiato per la colonna descrizione (matcha 'operazione', uno
    dei token di descrizione) PRIMA di arrivare alla vera colonna
    "Descrizione" - bug reale trovato sullo stesso export."""
    date_idx = next((i for i, h in enumerate(normalized_headers) if any(t in h for t in ['date', 'data'])), None)
    # Un rendiconto carta puo' avere DUE colonne "importo": quella nella
    # valuta originale dell'acquisto (vuota per le spese gia' in euro,
    # valorizzata solo per quelle in valuta estera) e quella vera nella
    # valuta del conto - bug reale trovato su un export reale, dove
    # "Importo originale in Divisa" veniva scelta perche' matcha 'import' per
    # prima, lasciando None (quindi riga scartata) su tutte le spese in euro
    # e facendo sopravvivere solo quelle in valuta estera. Scartiamo prima le
    # colonne "originale/divisa" e usiamo quelle solo se non c'e' alternativa.
    amount_candidates = [i for i, h in enumerate(normalized_headers) if any(t in h for t in ['amount', 'import', 'totale', 'valore'])]
    amount_idx = next((i for i in amount_candidates if not any(t in normalized_headers[i] for t in ['divisa', 'original'])), None)
    if amount_idx is None:
        amount_idx = amount_candidates[0] if amount_candidates else None
    debit_idx = next((i for i, h in enumerate(normalized_headers) if any(t in h for t in ['debit', 'addebito', 'uscit'])), None)
    credit_idx = next((i for i, h in enumerate(normalized_headers) if any(t in h for t in ['credit', 'accredito', 'entrat'])), None)
    taken = {i for i in (date_idx, amount_idx, debit_idx, credit_idx) if i is not None}
    # Priorita' a una colonna "descrizione completa/estesa" quando esiste
    # accanto a una "descrizione" generica: alcuni estratti (visto su un
    # export reale) hanno entrambe, dove quella generica e' solo il tipo di
    # movimento ("VISA DEBIT", "Pagamento Visa Debit") e quella completa ha il
    # vero esercente/causale - usare la prima trovata darebbe sempre la meno
    # utile delle due, dato che di norma la generica compare per prima.
    description_idx = next(
        (i for i, h in enumerate(normalized_headers)
         if i not in taken and any(t in h for t in ['completa', 'estesa', 'extended', 'complete'])),
        None,
    )
    if description_idx is None:
        description_idx = next(
            (i for i, h in enumerate(normalized_headers)
             if i not in taken and any(t in h for t in ['desc', 'causale', 'merchant', 'note', 'azienda', 'operazione'])),
        None,
    )
    return date_idx, amount_idx, debit_idx, credit_idx, description_idx


def parse_tabular_rows(rows: List[List[Any]]) -> List[Dict[str, Any]]:
    rows = [row for row in rows if any(cell is not None and str(cell).strip() != '' for cell in row)]
    if not rows:
        return []

    # Cerca la vera riga di intestazione invece di assumere che sia la prima:
    # la prima riga che, una volta normalizzata, individua sia una colonna
    # data sia una colonna importo/debito/credito. Due controlli di
    # plausibilita' in piu', altrimenti una riga di note discorsive nel
    # preambolo (vedi _HEADER_SCAN_ROWS) puo' combaciare per puro caso di
    # sottostringa - bug reale trovato su un export reale, dove una nota
    # legale che cita "carte di credito" e "data valuta" nello stesso
    # paragrafo veniva scambiata per intestazione:
    # 1) gli indici individuati devono essere tutti DISTINTI - un vero header
    #    non puo' avere la stessa cella che e' "la colonna data" E "la
    #    colonna credito" insieme, mentre una frase in un'unica cella si'
    #    (l'intera frase finisce nella stessa colonna 0);
    # 2) la riga deve avere piu' di una cella non vuota - un vero header
    #    elenca piu' nomi di colonna, una nota e' un'unica frase lunga in
    #    una sola cella con il resto della riga vuoto.
    header_row_idx = None
    date_idx = amount_idx = debit_idx = credit_idx = description_idx = None
    for i, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        non_empty_cells = sum(1 for cell in row if cell is not None and str(cell).strip() != '')
        if non_empty_cells < 2:
            continue
        normalized = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        d_idx, a_idx, deb_idx, cred_idx, desc_idx = _detect_columns(normalized)
        found = [idx for idx in (d_idx, a_idx, deb_idx, cred_idx) if idx is not None]
        if d_idx is not None and (a_idx is not None or deb_idx is not None or cred_idx is not None) and len(found) == len(set(found)):
            header_row_idx = i
            date_idx, amount_idx, debit_idx, credit_idx, description_idx = d_idx, a_idx, deb_idx, cred_idx, desc_idx
            break

    if header_row_idx is None:
        raise ValueError(
            "Intestazione non riconosciuta: non trovo una colonna data insieme a una colonna importo/entrate/uscite "
            f'nelle prime {_HEADER_SCAN_ROWS} righe del file.'
        )

    # Un rendiconto di carta di credito ha spesso UNA sola colonna importo,
    # sempre positiva (ogni riga e' un acquisto), senza alcun segno esplicito
    # ne' colonne addebiti/accrediti separate - bug reale trovato su un export
    # reale ("Movimenti Carta ****", "Carta di Credito" nel preambolo, colonna
    # unica "Importo (€)" sempre positiva anche per spese vere). Lo stesso
    # principio gia' usato per i PDF (vedi pdf_import.is_credit_card_statement)
    # qui si applica solo quando il segno arriva dalla colonna amount_idx
    # (non da un debit/credit gia' esplicito): solo se il valore e' positivo,
    # per non invertire un eventuale storno/rimborso gia' scritto negativo.
    preamble_text = ' '.join(
        str(cell).lower() for row in rows[:header_row_idx] for cell in row if cell is not None
    )
    is_credit_card_statement = bool(re.search(
        r'carta\s+di\s+credito|carta\s+revolving|movimenti\s+carta|limite\s+di\s+utilizzo', preamble_text,
    ))

    # Colonne diverse possono contenere una data valida per righe diverse
    # (es. "Data_Operazione" vale "-" per i movimenti carta non ancora
    # contabilizzati, dove solo "Data_Valuta" e' una data vera): proviamo
    # TUTTE le colonne che sembrano una data, non solo quella scelta come
    # riferimento principale, invece di scartare silenziosamente la riga.
    date_candidate_idxs = [date_idx] + [
        i for i, h in enumerate(
            [str(c).strip().lower() if c is not None else '' for c in rows[header_row_idx]]
        )
        if i != date_idx and any(t in h for t in ['date', 'data'])
    ]

    result = []
    for row in rows[header_row_idx + 1:]:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        date_text = None
        for idx in date_candidate_idxs:
            if idx >= len(row):
                continue
            value = row[idx]
            if value is None or str(value).strip() == '':
                continue
            try:
                if isinstance(value, date_cls):
                    # openpyxl restituisce gia' un datetime.date/datetime vero per
                    # le celle formattate come data in Excel: usarlo direttamente,
                    # SENZA passare da str(value) + dateutil, e' l'unico modo
                    # sicuro di evitare un bug reale trovato importando un vero
                    # estratto - str(datetime(2026,7,1)) produce "2026-07-01
                    # 00:00:00", e dateutil con dayfirst=True (necessario per le
                    # date scritte a mano in formato italiano nei CSV) scambia
                    # comunque giorno e mese quando il giorno reale e' <=12 anche
                    # se l'anno a 4 cifre rende la stringa gia' inequivocabile
                    # (risultato: 1 luglio letto come 7 gennaio). Le date scritte
                    # come testo (CSV, o celle Excel non formattate come data)
                    # restano sul percorso dateutil sotto, dove dayfirst serve
                    # davvero.
                    date_text = value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
                else:
                    date_text = parse_date(str(value), dayfirst=True).date().isoformat()
                break
            except Exception:
                continue
        if date_text is None:
            continue

        amount = None
        if amount_idx is not None and amount_idx < len(row):
            amount = parse_amount(row[amount_idx])
            if amount is not None and amount > 0 and is_credit_card_statement:
                amount = -amount
        if amount is None:
            debit = parse_amount(row[debit_idx]) if debit_idx is not None and debit_idx < len(row) else None
            credit = parse_amount(row[credit_idx]) if credit_idx is not None and credit_idx < len(row) else None
            if debit is not None and credit is None:
                amount = -abs(debit)
            elif credit is not None and debit is None:
                amount = abs(credit)
            elif debit is not None and credit is not None:
                amount = abs(credit) - abs(debit)
        if amount is None:
            continue

        description = ''
        if description_idx is not None and description_idx < len(row):
            description = str(row[description_idx] or '').strip()
        else:
            description = 'Importazione'

        result.append({
            'date': date_text,
            'amount': amount,
            'description': description,
        })
    return result


def ensure_int(value: Any) -> Optional[int]:
    if value is None or value == '':
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


@app.exception_handler(StarletteHTTPException)
async def custom_http_exception_handler(request: Request, exc: StarletteHTTPException):
    index_file = PUBLIC_DIR / 'index.html'
    if exc.status_code == 404 and not request.url.path.startswith('/api') and index_file.exists():
        return FileResponse(str(index_file), media_type='text/html')
    return JSONResponse({'error': exc.detail}, status_code=exc.status_code)


@app.get('/health')
def health():
    return {'status': 'ok', 'version': '1.0.0'}


@app.get('/api/backup/export')
def export_backup():
    wb = backup.build_backup_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"casaspese_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.post('/api/backup/import')
def import_backup(file: UploadFile = File(...)):
    data = file.file.read()
    try:
        summary = backup.import_backup_workbook(data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'File di backup non valido: {e}')
    return summary


@app.post('/api/admin/cleanup')
def cleanup():
    deleted_persons = execute("DELETE FROM persons WHERE name IS NULL OR trim(name)='' OR name='undefined'")
    deleted_accounts = execute("DELETE FROM accounts WHERE name IS NULL OR trim(name)='' OR name='undefined' OR name='..'")
    stats = fetchone(
        "SELECT (SELECT COUNT(*) FROM persons) AS persons, (SELECT COUNT(*) FROM accounts) AS accounts, "
        "(SELECT COUNT(*) FROM categories) AS categories, (SELECT COUNT(*) FROM transactions) AS transactions"
    )
    return {
        'deleted': {'persons': deleted_persons, 'accounts': deleted_accounts},
        'db': stats,
    }


@app.get('/api/setup/status')
def setup_status():
    setting = fetchone("SELECT value FROM settings WHERE key='setup_completed'")
    completed = json.loads(setting['value']) if setting else False
    person_count = fetchone('SELECT COUNT(*) AS c FROM persons')['c']
    account_count = fetchone('SELECT COUNT(*) AS c FROM accounts')['c']
    step = 1
    if person_count > 0:
        step = 2
    if account_count > 0:
        step = 3
    if completed:
        step = 4
    return {'completed': completed, 'step': step, 'personCount': person_count, 'accountCount': account_count}


@app.post('/api/setup/persons')
def setup_persons(payload: Dict[str, Any]):
    created = []
    for person in payload.get('persons', []):
        if person.get('name', '').strip():
            cursor = db.conn.execute(
                'INSERT INTO persons (name, email, color, is_primary) VALUES (?, ?, ?, ?)',
                (person['name'].strip(), person.get('email'), person.get('color', '#1D3557'), int(bool(person.get('isPrimary', False))))
            )
            db.conn.commit()
            created.append(_sanitize_person(fetchone('SELECT * FROM persons WHERE id = ?', (cursor.lastrowid,))))
    return JSONResponse(status_code=201, content=created)


@app.post('/api/setup/accounts')
def setup_accounts(payload: Dict[str, Any]):
    created = []
    for account in payload.get('accounts', []):
        if account.get('name', '').strip():
            cursor = db.conn.execute(
                'INSERT INTO accounts (name, bank, type, ownership, owner_id, co_owners, iban, color, nordigen_id, balance) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (
                    account['name'].strip(),
                    account.get('bank', 'other'),
                    account.get('type', 'checking'),
                    account.get('ownership', 'shared'),
                    ensure_int(account.get('ownerId')),
                    json.dumps(account.get('coOwners')) if account.get('coOwners') is not None else None,
                    account.get('iban'),
                    account.get('color'),
                    account.get('nordigenId'),
                    float(account['balance']) if account.get('balance') not in (None, '') else None,
                ),
            )
            db.conn.commit()
            created.append(fetchone('SELECT * FROM accounts WHERE id = ?', (cursor.lastrowid,)))
    return JSONResponse(status_code=201, content=created)


@app.post('/api/setup/categories')
def setup_categories(payload: Dict[str, Any]):
    budgets = payload.get('budgets', [])
    for budget in budgets:
        execute('UPDATE categories SET budget_monthly = ? WHERE id = ?', (budget.get('amount'), ensure_int(budget.get('categoryId'))))
    return {'updated': len(budgets)}


@app.post('/api/setup/complete')
def setup_complete(payload: Dict[str, Any]):
    if payload.get('aiProvider'):
        execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('ai_provider', json.dumps(payload['aiProvider'])))
    if payload.get('aiModel'):
        execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('ai_model', json.dumps(payload['aiModel'])))
    if payload.get('syncIntervalMinutes') is not None:
        execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('sync_interval_minutes', json.dumps(payload['syncIntervalMinutes'])))
    if payload.get('nordigenSecretId'):
        execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('nordigen_secret_id', json.dumps(payload['nordigenSecretId'])))
    if payload.get('nordigenSecretKey'):
        execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('nordigen_secret_key', json.dumps(payload['nordigenSecretKey'])))
    execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', ('setup_completed', json.dumps(True)))
    return {'completed': True}


def _sanitize_person(person: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Non restituire mai la password IMAP al client: la sostituisce con un
    booleano che indica solo se e' stata impostata."""
    if person is None:
        return None
    person = dict(person)
    person['imap_password_set'] = bool(person.get('imap_password'))
    person.pop('imap_password', None)
    return person


@app.get('/api/persons')
def list_persons():
    return [_sanitize_person(p) for p in fetchall('SELECT * FROM persons ORDER BY id')]


@app.get('/api/persons/{person_id}')
def get_person(person_id: int):
    person = fetchone('SELECT * FROM persons WHERE id = ?', (person_id,))
    if person is None:
        raise HTTPException(status_code=404, detail='Not found')
    return _sanitize_person(person)


@app.post('/api/persons')
def create_person(payload: Dict[str, Any]):
    if not payload.get('name', '').strip():
        raise HTTPException(status_code=400, detail='Nome obbligatorio')
    cursor = db.conn.execute(
        'INSERT INTO persons (name, email, color, is_primary, ha_user_id, imap_host, imap_port, imap_username, '
        'imap_password, imap_use_ssl, imap_folder) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (
            payload['name'].strip(),
            payload.get('email'),
            payload.get('color', '#1D3557'),
            int(bool(payload.get('isPrimary', False))),
            payload.get('haUserId'),
            payload.get('imapHost'),
            ensure_int(payload.get('imapPort')),
            payload.get('imapUsername'),
            payload.get('imapPassword'),
            int(bool(payload.get('imapUseSsl', True))),
            payload.get('imapFolder') or 'INBOX',
        ),
    )
    db.conn.commit()
    return JSONResponse(
        status_code=201,
        content=_sanitize_person(fetchone('SELECT * FROM persons WHERE id = ?', (cursor.lastrowid,))),
    )


@app.put('/api/persons/{person_id}')
def update_person(person_id: int, payload: Dict[str, Any]):
    if payload.get('name') is not None and not payload['name'].strip():
        raise HTTPException(status_code=400, detail='Nome obbligatorio')
    person = fetchone('SELECT * FROM persons WHERE id = ?', (person_id,))
    if person is None:
        raise HTTPException(status_code=404, detail='Not found')
    execute(
        'UPDATE persons SET name = ?, email = ?, color = ?, is_primary = ?, ha_user_id = ?, imap_host = ?, '
        'imap_port = ?, imap_username = ?, imap_password = ?, imap_use_ssl = ?, imap_folder = ? WHERE id = ?',
        (
            payload.get('name', person['name']).strip(),
            payload.get('email', person['email']),
            payload.get('color', person['color']),
            int(bool(payload.get('isPrimary', person['is_primary']))),
            payload.get('haUserId', person['ha_user_id']),
            payload.get('imapHost', person['imap_host']),
            ensure_int(payload['imapPort']) if 'imapPort' in payload else person['imap_port'],
            payload.get('imapUsername', person['imap_username']),
            payload.get('imapPassword', person['imap_password']),
            int(bool(payload.get('imapUseSsl', person['imap_use_ssl']))),
            payload.get('imapFolder', person['imap_folder']),
            person_id,
        ),
    )
    return _sanitize_person(fetchone('SELECT * FROM persons WHERE id = ?', (person_id,)))


@app.delete('/api/persons/{person_id}')
def delete_person(person_id: int):
    execute('DELETE FROM persons WHERE id = ?', (person_id,))
    return JSONResponse(status_code=204, content=None)


def _sanitize_mobile_token(token: Dict[str, Any]) -> Dict[str, Any]:
    return {k: v for k, v in token.items() if k != 'token_hash'}


@app.get('/api/mobile-tokens')
def list_mobile_tokens(request: Request):
    params = request.query_params
    sql = 'SELECT * FROM mobile_tokens'
    args: tuple = ()
    if person_id := ensure_int(params.get('personId')):
        sql += ' WHERE person_id = ?'
        args = (person_id,)
    sql += ' ORDER BY id DESC'
    return [_sanitize_mobile_token(t) for t in fetchall(sql, args)]


@app.post('/api/mobile-tokens')
def create_mobile_token(payload: Dict[str, Any]):
    person_id = ensure_int(payload.get('personId'))
    if not person_id:
        raise HTTPException(status_code=400, detail='personId obbligatorio')
    person = fetchone('SELECT * FROM persons WHERE id = ?', (person_id,))
    if person is None:
        raise HTTPException(status_code=404, detail='Persona non trovata')
    raw_token = access.generate_mobile_token()
    cursor = db.conn.execute(
        'INSERT INTO mobile_tokens (person_id, token_hash, label) VALUES (?, ?, ?)',
        (person_id, access.hash_mobile_token(raw_token), payload.get('label')),
    )
    db.conn.commit()
    if not config.PUBLIC_URL:
        url = None
    else:
        url = f'{config.PUBLIC_URL}/#/mobile/link?token={raw_token}'
    return JSONResponse(
        status_code=201,
        content={
            **_sanitize_mobile_token(fetchone('SELECT * FROM mobile_tokens WHERE id = ?', (cursor.lastrowid,))),
            'token': raw_token,
            'url': url,
        },
    )


@app.delete('/api/mobile-tokens/{token_id}')
def revoke_mobile_token(token_id: int):
    execute("UPDATE mobile_tokens SET revoked_at = datetime('now') WHERE id = ? AND revoked_at IS NULL", (token_id,))
    return JSONResponse(status_code=204, content=None)


@app.get('/api/mobile/me')
def mobile_me(request: Request):
    """Chi sta usando la PWA in questo momento (via token Bearer) - usata dalla
    schermata di scansione scontrino per salutare l'utente e preselezionare la
    persona come pagatore, senza dover ripetere la risoluzione lato client."""
    person = access.get_current_person(request)
    if person is None:
        raise HTTPException(status_code=401, detail='Nessuna persona riconosciuta')
    return _sanitize_person(person)


@app.post('/api/persons/{person_id}/email-backfill')
def email_backfill_endpoint(person_id: int, payload: Dict[str, Any]):
    person = fetchone('SELECT * FROM persons WHERE id = ?', (person_id,))
    if person is None:
        raise HTTPException(status_code=404, detail='Not found')
    senders = payload.get('senders') or ['paypal.com', 'amazon.it', 'amazon.com']
    subject_keywords = payload.get('subjectKeywords') or email_backfill.DEFAULT_SUBJECT_KEYWORDS
    try:
        result = email_backfill.run_backfill(
            person,
            senders=senders,
            date_from=payload.get('dateFrom'),
            date_to=payload.get('dateTo'),
            subject_keywords=subject_keywords,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return result


@app.post('/api/persons/{person_id}/email-backfill-stream')
def email_backfill_stream(person_id: int, payload: Dict[str, Any]):
    """Come /email-backfill ma risponde con un flusso SSE (stage/progress/done)
    invece di un'unica risposta: la scansione IMAP di una casella con molte mail
    puo' richiedere svariati secondi e senza feedback il frontend sembra bloccato."""
    person = fetchone('SELECT * FROM persons WHERE id = ?', (person_id,))
    if person is None:
        raise HTTPException(status_code=404, detail='Not found')
    senders = payload.get('senders') or ['paypal.com', 'amazon.it', 'amazon.com']
    subject_keywords = payload.get('subjectKeywords') or email_backfill.DEFAULT_SUBJECT_KEYWORDS

    def sse(event: str, data: Dict[str, Any]) -> str:
        return f'event: {event}\ndata: {json.dumps(data)}\n\n'

    def event_stream():
        try:
            for update in email_backfill.run_backfill_iter(
                person,
                senders=senders,
                date_from=payload.get('dateFrom'),
                date_to=payload.get('dateTo'),
                subject_keywords=subject_keywords,
            ):
                if update.get('done'):
                    yield sse('done', update)
                else:
                    yield sse('progress', update)
        except ValueError as e:
            yield sse('error', {'detail': str(e)})

    return StreamingResponse(event_stream(), media_type='text/event-stream')


@app.post('/api/persons/{person_id}/email-poll-now')
def email_poll_now(person_id: int):
    """Forza subito un controllo IMAP incrementale (solo mail nuove, vedi
    email_backfill.run_incremental_poll) invece di aspettare il prossimo giro
    automatico di email_poller - utile appena configurate le credenziali per
    verificare che funzionino senza aspettare fino a sync_interval_minutes."""
    person = fetchone('SELECT * FROM persons WHERE id = ?', (person_id,))
    if person is None:
        raise HTTPException(status_code=404, detail='Not found')
    try:
        result = email_backfill.run_incremental_poll(
            person, senders=email_poller.DEFAULT_SENDERS, subject_keywords=email_backfill.DEFAULT_SUBJECT_KEYWORDS,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    execute(
        "UPDATE persons SET imap_last_uid = ?, imap_uidvalidity = ?, imap_last_checked_at = datetime('now') WHERE id = ?",
        (result['newLastUid'], result['newUidValidity'], person_id),
    )
    return result


def _opening_balance_category_id() -> Optional[int]:
    row = fetchone("SELECT id FROM categories WHERE code = 'SALDO_INIT'")
    return row['id'] if row else None


def _compute_account_balances() -> Dict[int, float]:
    """Saldo di ogni conto = importo dell'ultimo checkpoint 'saldo iniziale'
    (per data, vedi POST /api/accounts/{id}/opening-balance) + la somma di
    tutti i movimenti (esclusi altri checkpoint) datati dopo quel checkpoint.
    Un conto senza alcun checkpoint somma semplicemente tutto lo storico
    (comportamento invariato per i conti creati prima di questa funzionalita').
    Due query per TUTTI i conti invece di un ciclo per-conto."""
    cat_id = _opening_balance_category_id()
    if cat_id is None:
        return {}
    checkpoints = fetchall(
        '''SELECT account_id, cp_date, cp_amount FROM (
             SELECT account_id, date AS cp_date, amount AS cp_amount,
                    ROW_NUMBER() OVER (PARTITION BY account_id ORDER BY date DESC, id DESC) AS rn
             FROM transactions WHERE category_id = ?
           ) WHERE rn = 1''',
        (cat_id,),
    )
    cp_by_account = {row['account_id']: (row['cp_date'], row['cp_amount']) for row in checkpoints}
    sums_after = fetchall(
        '''SELECT t.account_id AS account_id, COALESCE(SUM(t.amount),0) AS total
           FROM transactions t
           LEFT JOIN (
             SELECT account_id, date AS cp_date,
                    ROW_NUMBER() OVER (PARTITION BY account_id ORDER BY date DESC, id DESC) AS rn
             FROM transactions WHERE category_id = ?
           ) cp ON cp.account_id = t.account_id AND cp.rn = 1
           WHERE (t.category_id IS NULL OR t.category_id != ?) AND t.date > COALESCE(cp.cp_date, '0000-00-00')
           GROUP BY t.account_id''',
        (cat_id, cat_id),
    )
    sum_by_account = {row['account_id']: row['total'] for row in sums_after}
    account_ids = set(cp_by_account) | set(sum_by_account)
    return {
        account_id: cp_by_account.get(account_id, (None, 0.0))[1] + sum_by_account.get(account_id, 0.0)
        for account_id in account_ids
    }


def _with_computed_balance(row: Optional[Dict[str, Any]], balances: Optional[Dict[int, float]] = None) -> Optional[Dict[str, Any]]:
    if row is None:
        return None
    if balances is None:
        balances = _compute_account_balances()
    row['balance'] = round(balances.get(row['id'], 0.0), 2)
    return row


@app.get('/api/accounts')
def list_accounts(request: Request):
    current_person = access.get_current_person(request)
    vis_clause, vis_args = access.account_visibility(current_person)
    rows = fetchall(f'SELECT * FROM accounts WHERE is_active = 1 AND {vis_clause} ORDER BY id', vis_args)
    balances = _compute_account_balances()
    return [_with_computed_balance(row, balances) for row in rows]


@app.get('/api/accounts/{account_id}')
def get_account(account_id: int, request: Request):
    account = fetchone('SELECT * FROM accounts WHERE id = ?', (account_id,))
    if account is None or not access.can_see_account(account, access.get_current_person(request)):
        raise HTTPException(status_code=404, detail='Not found')
    return _with_computed_balance(account)


@app.post('/api/accounts')
def create_account(payload: Dict[str, Any]):
    if not payload.get('name', '').strip():
        raise HTTPException(status_code=400, detail='Nome obbligatorio')
    cursor = db.conn.execute(
        'INSERT INTO accounts (name, bank, type, ownership, owner_id, co_owners, iban, color, nordigen_id, balance, settlement_account_id, card_number) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (
            payload['name'].strip(),
            payload.get('bank', 'other'),
            payload.get('type', 'checking'),
            payload.get('ownership', 'shared'),
            ensure_int(payload.get('ownerId')),
            json.dumps(payload.get('coOwners')) if payload.get('coOwners') is not None else None,
            payload.get('iban'),
            payload.get('color'),
            payload.get('nordigenId'),
            float(payload['balance']) if payload.get('balance') not in (None, '') else None,
            ensure_int(payload.get('settlementAccountId')),
            payload.get('cardNumber'),
        ),
    )
    db.conn.commit()
    created = fetchone('SELECT * FROM accounts WHERE id = ?', (cursor.lastrowid,))
    return JSONResponse(status_code=201, content=_with_computed_balance(created))


@app.put('/api/accounts/{account_id}')
def update_account(account_id: int, payload: Dict[str, Any]):
    account = fetchone('SELECT * FROM accounts WHERE id = ?', (account_id,))
    if account is None:
        raise HTTPException(status_code=404, detail='Not found')
    if payload.get('name') is not None and not payload['name'].strip():
        raise HTTPException(status_code=400, detail='Nome obbligatorio')
    execute(
        'UPDATE accounts SET name = ?, bank = ?, type = ?, ownership = ?, owner_id = ?, iban = ?, color = ?, balance = ?, is_active = ?, settlement_account_id = ?, card_number = ? WHERE id = ?',
        (
            payload.get('name', account['name']).strip(),
            payload.get('bank', account['bank']),
            payload.get('type', account['type']),
            payload.get('ownership', account['ownership']),
            ensure_int(payload['ownerId']) if 'ownerId' in payload else account['owner_id'],
            payload.get('iban', account['iban']),
            payload.get('color', account['color']),
            float(payload['balance']) if payload.get('balance') not in (None, '') else None,
            int(bool(payload.get('isActive', account['is_active']))),
            ensure_int(payload['settlementAccountId']) if 'settlementAccountId' in payload else account['settlement_account_id'],
            payload.get('cardNumber', account['card_number']),
            account_id,
        ),
    )
    return _with_computed_balance(fetchone('SELECT * FROM accounts WHERE id = ?', (account_id,)))


@app.delete('/api/accounts/{account_id}')
def delete_account(account_id: int):
    execute('DELETE FROM accounts WHERE id = ?', (account_id,))
    return JSONResponse(status_code=204, content=None)


@app.get('/api/accounts/{account_id}/opening-balance')
def list_opening_balances(account_id: int):
    cat_id = _opening_balance_category_id()
    if cat_id is None:
        return []
    return fetchall(
        'SELECT * FROM transactions WHERE account_id = ? AND category_id = ? ORDER BY date DESC',
        (account_id, cat_id),
    )


@app.post('/api/accounts/{account_id}/opening-balance')
def set_opening_balance(account_id: int, payload: Dict[str, Any]):
    """Crea o aggiorna (upsert per anno solare) il checkpoint 'saldo iniziale'
    di un conto: vedi _compute_account_balances per come viene poi usato nel
    calcolo del saldo mostrato all'utente."""
    account = fetchone('SELECT * FROM accounts WHERE id = ?', (account_id,))
    if account is None:
        raise HTTPException(status_code=404, detail='Not found')
    date = payload.get('date')
    if not date:
        raise HTTPException(status_code=400, detail='Data obbligatoria')
    try:
        amount = float(payload['amount'])
    except (KeyError, TypeError, ValueError):
        raise HTTPException(status_code=400, detail='Importo non valido')
    cat_id = _opening_balance_category_id()
    if cat_id is None:
        raise HTTPException(status_code=500, detail="Categoria di sistema 'Saldo iniziale' non trovata")
    year = date[:4]
    is_personal_account = account['ownership'] == 'personal'
    destination = 'personal' if is_personal_account else 'family'
    paid_by_person_id = account['owner_id'] if is_personal_account else None
    existing = fetchone(
        "SELECT id FROM transactions WHERE account_id = ? AND category_id = ? AND strftime('%Y', date) = ?",
        (account_id, cat_id, year),
    )
    if existing:
        execute(
            "UPDATE transactions SET date = ?, amount = ?, updated_at = (datetime('now')) WHERE id = ?",
            (date, amount, existing['id']),
        )
        transaction_id = existing['id']
    else:
        cursor = db.conn.execute(
            'INSERT INTO transactions (date, amount, description_raw, merchant_name, category_id, account_id, '
            'destination, paid_by_person_id, is_confirmed, import_source) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)',
            (date, amount, 'Saldo iniziale', 'Saldo iniziale', cat_id, account_id, destination, paid_by_person_id, 'manual'),
        )
        db.conn.commit()
        transaction_id = cursor.lastrowid
    return fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))


@app.get('/api/categories')
def list_categories():
    return fetchall('SELECT * FROM categories ORDER BY sort_order')


@app.get('/api/categories/defaults')
def list_category_defaults():
    return fetchall('SELECT * FROM categories ORDER BY sort_order')


@app.get('/api/categories/{category_id}')
def get_category(category_id: int):
    category = fetchone('SELECT * FROM categories WHERE id = ?', (category_id,))
    if category is None:
        raise HTTPException(status_code=404, detail='Not found')
    return category


def _validate_category_parent(category_id: Optional[int], parent_id: Optional[int]) -> None:
    """La gerarchia e' volutamente limitata a 2 livelli (categoria -> sotto-
    categoria), per tenere il modello semplice: niente sotto-categorie di
    sotto-categorie. Vedi Categories.vue per la UI ad albero costruita su
    questo stesso vincolo."""
    if parent_id is None:
        return
    if parent_id == category_id:
        raise HTTPException(status_code=400, detail='Una categoria non puo\' essere genitore di se stessa')
    parent = fetchone('SELECT id, parent_id FROM categories WHERE id = ?', (parent_id,))
    if parent is None:
        raise HTTPException(status_code=400, detail='Categoria padre non trovata')
    if parent['parent_id'] is not None:
        raise HTTPException(status_code=400, detail='Una sotto-categoria non puo\' avere a sua volta sotto-categorie')
    if category_id is not None:
        has_children = fetchone('SELECT id FROM categories WHERE parent_id = ? LIMIT 1', (category_id,))
        if has_children is not None:
            raise HTTPException(status_code=400, detail='Questa categoria ha gia\' delle sotto-categorie: spostale o rimuovile prima di assegnarle un genitore')


@app.post('/api/categories')
def create_category(payload: Dict[str, Any]):
    if not payload.get('name', '').strip():
        raise HTTPException(status_code=400, detail='Nome obbligatorio')
    parent_id = ensure_int(payload.get('parentId'))
    _validate_category_parent(None, parent_id)
    cursor = db.conn.execute(
        'INSERT INTO categories (code, name, icon, color, type, budget_monthly, budget_annual, ai_keywords, parent_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (
            payload.get('code'),
            payload['name'].strip(),
            payload.get('icon'),
            payload.get('color'),
            payload.get('type', 'expense'),
            float(payload['budgetMonthly']) if payload.get('budgetMonthly') not in (None, '') else None,
            float(payload['budgetAnnual']) if payload.get('budgetAnnual') not in (None, '') else None,
            payload.get('aiKeywords'),
            parent_id,
        ),
    )
    db.conn.commit()
    return JSONResponse(status_code=201, content=fetchone('SELECT * FROM categories WHERE id = ?', (cursor.lastrowid,)))


@app.put('/api/categories/{category_id}')
def update_category(category_id: int, payload: Dict[str, Any]):
    category = fetchone('SELECT * FROM categories WHERE id = ?', (category_id,))
    if category is None:
        raise HTTPException(status_code=404, detail='Not found')
    if payload.get('name') is not None and not payload['name'].strip():
        raise HTTPException(status_code=400, detail='Nome obbligatorio')

    def budget_value(key, current):
        if key not in payload:
            return current
        return float(payload[key]) if payload[key] not in (None, '') else None

    parent_id = ensure_int(payload['parentId']) if 'parentId' in payload else category['parent_id']
    if parent_id != category['parent_id']:
        _validate_category_parent(category_id, parent_id)

    execute(
        'UPDATE categories SET code = ?, name = ?, icon = ?, color = ?, type = ?, budget_monthly = ?, budget_annual = ?, is_active = ?, ai_keywords = ?, parent_id = ? WHERE id = ?',
        (
            payload.get('code', category['code']),
            payload.get('name', category['name']).strip() if payload.get('name') is not None else category['name'],
            payload.get('icon', category['icon']),
            payload.get('color', category['color']),
            payload.get('type', category['type']),
            budget_value('budgetMonthly', category['budget_monthly']),
            budget_value('budgetAnnual', category['budget_annual']),
            int(bool(payload.get('isActive', category['is_active']))),
            payload.get('aiKeywords', category['ai_keywords']),
            parent_id,
            category_id,
        ),
    )
    return fetchone('SELECT * FROM categories WHERE id = ?', (category_id,))


@app.delete('/api/categories/{category_id}')
def delete_category(category_id: int):
    """Elimina davvero la categoria (non solo un disattiva): le transazioni che
    la usavano restano, ma senza categoria (da ricategorizzare manualmente).
    Le sotto-categorie restano invece invariate (non cancellate a cascata), ma
    tornano categorie di primo livello, altrimenti resterebbero con un
    parent_id orfano che punta a una riga non piu' esistente."""
    execute('UPDATE categories SET parent_id = NULL WHERE parent_id = ?', (category_id,))
    execute('UPDATE transactions SET category_id = NULL WHERE category_id = ?', (category_id,))
    execute('UPDATE transactions SET ai_category_id = NULL WHERE ai_category_id = ?', (category_id,))
    execute('DELETE FROM budgets WHERE category_id = ?', (category_id,))
    execute('DELETE FROM categories WHERE id = ?', (category_id,))
    return JSONResponse(status_code=204, content=None)


@app.get('/api/transactions')
def list_transactions(request: Request):
    params = request.query_params
    current_person = access.get_current_person(request)
    filters = []
    sql = (
        'SELECT transactions.*, '
        '(SELECT COUNT(*) FROM documents WHERE documents.transaction_id = transactions.id) AS attachment_count, '
        '(SELECT id FROM email_receipts WHERE email_receipts.matched_transaction_id = transactions.id LIMIT 1) AS email_receipt_id '
        'FROM transactions'
    )
    args: List[Any] = []
    if month := params.get('month'):
        filters.append('date LIKE ?')
        args.append(f'{month}%')
    if account_id := ensure_int(params.get('accountId')):
        filters.append('account_id = ?')
        args.append(account_id)
    if category_id := ensure_int(params.get('categoryId')):
        filters.append('category_id = ?')
        args.append(category_id)
    if destination := params.get('destination'):
        filters.append('destination = ?')
        args.append(destination)
    if person_id := ensure_int(params.get('personId')):
        filters.append('paid_by_person_id = ?')
        args.append(person_id)
    if params.get('unconfirmed') == 'true':
        filters.append('is_confirmed = 0')
    elif params.get('confirmed') == 'true':
        filters.append('is_confirmed = 1')
    if reimbursable := params.get('reimbursable'):
        if reimbursable == 'pending':
            filters.append('is_reimbursable = 1 AND reimbursed_at IS NULL')
        elif reimbursable == 'reimbursed':
            filters.append('is_reimbursable = 1 AND reimbursed_at IS NOT NULL')
        elif reimbursable == 'all':
            filters.append('is_reimbursable = 1')
    vis_clause, vis_args = access.transaction_visibility(current_person)
    filters.append(vis_clause)
    args.extend(vis_args)
    if filters:
        sql += ' WHERE ' + ' AND '.join(filters)
    sql += ' ORDER BY date DESC'
    limit = ensure_int(params.get('limit')) or 200
    sql += ' LIMIT ?'
    args.append(limit)
    if offset := ensure_int(params.get('offset')):
        sql += ' OFFSET ?'
        args.append(offset)
    return fetchall(sql, tuple(args))


@app.get('/api/transactions/pending-ai')
def pending_ai(request: Request):
    current_person = access.get_current_person(request)
    vis_clause, vis_args = access.transaction_visibility(current_person)
    return fetchall(
        f'SELECT * FROM transactions WHERE is_confirmed = 0 AND ai_category_id IS NOT NULL AND {vis_clause} '
        'ORDER BY date DESC LIMIT 100',
        vis_args,
    )


@app.get('/api/transactions/{transaction_id}')
def get_transaction(transaction_id: int, request: Request):
    tx = fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))
    if tx is None or not access.can_see_transaction(tx, access.get_current_person(request)):
        raise HTTPException(status_code=404, detail='Not found')
    return tx


@app.post('/api/transactions')
def create_transaction(payload: Dict[str, Any]):
    if not payload.get('date') or payload.get('amount') is None or not payload.get('description') or not ensure_int(payload.get('accountId')):
        raise HTTPException(status_code=400, detail='Campi obbligatori mancanti')
    is_reimbursable = bool(payload.get('isReimbursable', False))
    cursor = db.conn.execute(
        'INSERT INTO transactions (date, amount, currency, description_raw, merchant_name, category_id, account_id, '
        'destination, paid_by_person_id, split_person_id, split_ratio, is_cash, is_confirmed, import_source, notes, '
        'is_reimbursable, reimbursement_amount) '
        'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (
            payload['date'],
            float(payload['amount']),
            payload.get('currency', 'EUR'),
            payload['description'],
            payload.get('merchantName') or payload['description'],
            ensure_int(payload.get('categoryId')),
            ensure_int(payload.get('accountId')),
            payload.get('destination', 'family'),
            ensure_int(payload.get('paidByPersonId')),
            ensure_int(payload.get('splitPersonId')),
            float(payload['splitRatio']) if payload.get('splitRatio') not in (None, '') else 0.5,
            int(bool(payload.get('isCash', False))),
            1,
            'manual',
            payload.get('notes'),
            int(is_reimbursable),
            float(payload['reimbursementAmount']) if is_reimbursable and payload.get('reimbursementAmount') not in (None, '') else None,
        ),
    )
    db.conn.commit()
    return JSONResponse(status_code=201, content=fetchone('SELECT * FROM transactions WHERE id = ?', (cursor.lastrowid,)))


@app.post('/api/transactions/ai-parse')
def ai_parse_transaction(payload: Dict[str, Any]):
    """Trasforma una spesa descritta in linguaggio naturale (es. '23€ pizza ieri
    sera con amex') in una bozza da precompilare nel form di inserimento
    manuale: l'utente la rivede e la conferma sempre a mano (stesso principio
    delle categorie suggerite dall'AI sugli import, mai un inserimento alla
    cieca) - questo endpoint non scrive nulla su transactions."""
    text = (payload.get('text') or '').strip()
    if not text:
        raise HTTPException(status_code=400, detail='Testo mancante')

    categories = fetchall("SELECT id, name FROM categories WHERE is_active = 1 AND type = 'expense' ORDER BY sort_order")
    accounts = fetchall('SELECT id, name FROM accounts WHERE is_active = 1')
    today = datetime.utcnow().strftime('%Y-%m-%d')
    cats_text = '\n'.join(f"- id={c['id']}: {c['name']}" for c in categories)
    accs_text = '\n'.join(f"- id={a['id']}: {a['name']}" for a in accounts)
    prompt = f"""Sei un assistente che trasforma la descrizione informale di una spesa in dati strutturati, per una famiglia italiana.

Data di oggi: {today}

Categorie disponibili (usa solo questi id, o null se nessuna e' plausibile):
{cats_text}

Conti disponibili (usa solo questi id, o null se il conto non e' menzionato o non lo riconosci):
{accs_text}

Testo della spesa: "{text}"

Rispondi SOLO con un oggetto JSON valido (nessun testo extra, nessun blocco markdown):
{{"amount": 23.50, "description": "descrizione sintetica della spesa", "date": "YYYY-MM-DD", "categoryId": 12, "accountId": 3}}

Regole:
- amount sempre un numero positivo
- date: risolvi riferimenti relativi ("ieri", "oggi", "lunedi' scorso") rispetto alla data di oggi indicata sopra, formato YYYY-MM-DD
- se un campo non e' determinabile con sicurezza, usa null per quel campo"""

    try:
        content = ai_client.ask_ai(prompt, task_name='casaspese_quick_add', max_tokens=300)
        data = ai_client.parse_json_object(content)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    return {
        'amount': data.get('amount'),
        'description': data.get('description'),
        'date': data.get('date'),
        'categoryId': ensure_int(data.get('categoryId')),
        'accountId': ensure_int(data.get('accountId')),
    }


@app.post('/api/transactions/ai-parse-receipt')
async def ai_parse_receipt(file: UploadFile = File(...)):
    """Come /api/transactions/ai-parse ma partendo da una foto di scontrino
    invece che da un testo libero: usata dalla schermata mobile di scansione.
    Come ai-parse, non scrive nulla - restituisce solo una bozza da rivedere."""
    image_bytes = await file.read()
    if not image_bytes:
        raise HTTPException(status_code=400, detail='Immagine mancante')

    categories = fetchall("SELECT id, name FROM categories WHERE is_active = 1 AND type = 'expense' ORDER BY sort_order")
    today = datetime.utcnow().strftime('%Y-%m-%d')
    cats_text = '\n'.join(f"- id={c['id']}: {c['name']}" for c in categories)
    prompt = f"""Sei un assistente che legge una foto di uno scontrino/ricevuta italiana e ne estrae i dati.

Data di oggi (usala solo se sullo scontrino non compare una data leggibile): {today}

Categorie disponibili (usa solo questi id, o null se nessuna e' plausibile):
{cats_text}

Rispondi SOLO con un oggetto JSON valido (nessun testo extra, nessun blocco markdown):
{{"amount": 23.50, "merchantName": "nome esercente", "date": "YYYY-MM-DD", "categoryId": 12}}

Regole:
- amount e' il totale pagato (numero positivo)
- date nel formato YYYY-MM-DD
- se un campo non e' leggibile con sicurezza, usa null per quel campo"""

    try:
        content = ai_client.ask_ai_with_image(prompt, image_bytes, file.filename or 'scontrino.jpg')
        data = ai_client.parse_json_object(content)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    return {
        'amount': data.get('amount'),
        'merchantName': data.get('merchantName'),
        'date': data.get('date'),
        'categoryId': ensure_int(data.get('categoryId')),
    }


@app.put('/api/transactions/{transaction_id}')
def update_transaction(transaction_id: int, payload: Dict[str, Any], request: Request):
    tx = fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))
    if tx is None or not access.can_see_transaction(tx, access.get_current_person(request)):
        raise HTTPException(status_code=404, detail='Not found')
    # Il form di modifica non ha un campo merchantName separato: se l'utente
    # cambia la descrizione, aggiorna anche il nome visualizzato (merchant_name),
    # a meno che non sia gia' stato arricchito esplicitamente via merchantName.
    merchant_name = payload.get('merchantName', payload.get('description', tx['merchant_name']))
    is_reimbursable = bool(payload.get('isReimbursable', tx['is_reimbursable']))
    if not is_reimbursable:
        reimbursement_amount = None
    elif 'reimbursementAmount' in payload:
        reimbursement_amount = float(payload['reimbursementAmount']) if payload.get('reimbursementAmount') not in (None, '') else None
    else:
        reimbursement_amount = tx['reimbursement_amount']
    execute(
        'UPDATE transactions SET date = ?, amount = ?, description_raw = ?, merchant_name = ?, category_id = ?, '
        'account_id = ?, destination = ?, paid_by_person_id = ?, split_person_id = ?, split_ratio = ?, is_cash = ?, '
        "is_confirmed = ?, notes = ?, is_reimbursable = ?, reimbursement_amount = ?, updated_at = (datetime('now')) WHERE id = ?",
        (
            payload.get('date', tx['date']),
            float(payload['amount']) if payload.get('amount') not in (None, '') else tx['amount'],
            payload.get('description', tx['description_raw']),
            merchant_name,
            ensure_int(payload['categoryId']) if 'categoryId' in payload else tx['category_id'],
            ensure_int(payload['accountId']) if 'accountId' in payload else tx['account_id'],
            payload.get('destination', tx['destination']),
            ensure_int(payload['paidByPersonId']) if 'paidByPersonId' in payload else tx['paid_by_person_id'],
            ensure_int(payload['splitPersonId']) if 'splitPersonId' in payload else tx['split_person_id'],
            float(payload['splitRatio']) if payload.get('splitRatio') not in (None, '') else tx['split_ratio'],
            int(bool(payload.get('isCash', tx['is_cash']))),
            int(bool(payload.get('isConfirmed', tx['is_confirmed']))),
            payload.get('notes', tx['notes']),
            int(is_reimbursable),
            reimbursement_amount,
            transaction_id,
        ),
    )
    return fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))


@app.post('/api/transactions/{transaction_id}/toggle-reimbursed')
def toggle_reimbursed(transaction_id: int, request: Request):
    """Segna/riapre una spesa 'da rimborsare' quando l'azienda accredita (o
    smentisce) il rimborso. reimbursed_at NULL = ancora in attesa."""
    tx = fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))
    if tx is None or not access.can_see_transaction(tx, access.get_current_person(request)):
        raise HTTPException(status_code=404, detail='Not found')
    if not tx['is_reimbursable']:
        raise HTTPException(status_code=400, detail="La transazione non e' marcata come da rimborsare")
    new_value = None if tx['reimbursed_at'] else db.conn.execute("SELECT datetime('now')").fetchone()[0]
    execute('UPDATE transactions SET reimbursed_at = ? WHERE id = ?', (new_value, transaction_id))
    return fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))


@app.delete('/api/transactions/{transaction_id}')
def delete_transaction(transaction_id: int, request: Request):
    tx = fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))
    if tx is None or not access.can_see_transaction(tx, access.get_current_person(request)):
        raise HTTPException(status_code=404, detail='Not found')
    execute('DELETE FROM transactions WHERE id = ?', (transaction_id,))
    return JSONResponse(status_code=204, content=None)


@app.post('/api/transactions/confirm-bulk')
def confirm_bulk(payload: Dict[str, Any], request: Request):
    ids = [ensure_int(x) for x in payload.get('ids', []) if ensure_int(x) is not None]
    current_person = access.get_current_person(request)
    confirmed = 0
    for tx_id in ids:
        tx = fetchone('SELECT * FROM transactions WHERE id = ?', (tx_id,))
        if tx is None or not access.can_see_transaction(tx, current_person):
            continue
        if tx['ai_category_id'] is not None:
            execute('UPDATE transactions SET category_id = ?, is_confirmed = 1 WHERE id = ?', (tx['ai_category_id'], tx_id))
            confirmed += 1
    return {'confirmed': confirmed}


@app.post('/api/transactions/reject-ai-bulk')
def reject_ai_bulk(payload: Dict[str, Any], request: Request):
    """Scarta il suggerimento AI (categoria + confidenza) senza confermare ne'
    eliminare la transazione: torna semplicemente 'da categorizzare a mano'."""
    ids = [ensure_int(x) for x in payload.get('ids', []) if ensure_int(x) is not None]
    current_person = access.get_current_person(request)
    rejected = 0
    for tx_id in ids:
        tx = fetchone('SELECT * FROM transactions WHERE id = ?', (tx_id,))
        if tx is None or tx['is_confirmed'] or not access.can_see_transaction(tx, current_person):
            continue
        execute('UPDATE transactions SET ai_category_id = NULL, ai_confidence = NULL WHERE id = ?', (tx_id,))
        rejected += 1
    return {'rejected': rejected}


@app.post('/api/transactions/bulk-update')
def bulk_update_transactions(payload: Dict[str, Any], request: Request):
    """Sposta in blocco una lista di transazioni su categoria/conto/destinazione/
    pagato da/stato conferma. Applica solo i campi presenti nel payload,
    lasciando invariati gli altri (stesso comportamento di update_transaction)."""
    ids = [ensure_int(x) for x in payload.get('ids', []) if ensure_int(x) is not None]
    if not ids:
        return {'updated': 0, 'skipped': 0}
    current_person = access.get_current_person(request)
    updated = 0
    for tx_id in ids:
        tx = fetchone('SELECT * FROM transactions WHERE id = ?', (tx_id,))
        if tx is None or not access.can_see_transaction(tx, current_person):
            continue
        if 'categoryId' in payload:
            category_id = ensure_int(payload['categoryId'])
        elif payload.get('isConfirmed') and tx['category_id'] is None and tx['ai_category_id'] is not None:
            # Confermare in blocco senza scegliere esplicitamente una categoria
            # (es. pulsante "Conferma" della barra di selezione) deve accettare
            # la proposta AI come per la conferma singola/da banner, non
            # lasciare la transazione "confermata" ma senza categoria - bug
            # reale: prima qui restava category_id NULL nonostante ai_category_id
            # gia' presente.
            category_id = tx['ai_category_id']
        else:
            category_id = tx['category_id']
        execute(
            'UPDATE transactions SET category_id = ?, account_id = ?, destination = ?, paid_by_person_id = ?, '
            "is_confirmed = ?, is_reimbursable = ?, updated_at = (datetime('now')) WHERE id = ?",
            (
                category_id,
                ensure_int(payload['accountId']) if 'accountId' in payload else tx['account_id'],
                payload.get('destination', tx['destination']),
                ensure_int(payload['paidByPersonId']) if 'paidByPersonId' in payload else tx['paid_by_person_id'],
                int(bool(payload.get('isConfirmed', tx['is_confirmed']))),
                int(bool(payload.get('isReimbursable', tx['is_reimbursable']))),
                tx_id,
            ),
        )
        updated += 1
    return {'updated': updated, 'skipped': len(ids) - updated}


@app.post('/api/transactions/categorize-ai')
def categorize_ai_bulk(payload: Dict[str, Any]):
    """Rilancia a mano il riconoscimento AI della categoria su una lista di
    transazioni scelte dall'utente in un momento qualunque (non legate a un
    import appena fatto) - es. spese inserite manualmente o importate prima
    che questa categoria esistesse. Vedi categorize.categorize_selected per
    la logica (salta quelle gia' categorizzate/con suggerimento pendente)."""
    ids = [ensure_int(x) for x in payload.get('ids', []) if ensure_int(x) is not None]
    return categorize.categorize_selected(ids)


@app.post('/api/transactions/bulk-delete')
def bulk_delete_transactions(payload: Dict[str, Any], request: Request):
    """Elimina in blocco una lista di transazioni, es. per ripulire un import
    sbagliato (conto/segno errato) prima di reimportare."""
    ids = [ensure_int(x) for x in payload.get('ids', []) if ensure_int(x) is not None]
    if not ids:
        return {'deleted': 0, 'skipped': 0}
    current_person = access.get_current_person(request)
    deleted = 0
    for tx_id in ids:
        tx = fetchone('SELECT * FROM transactions WHERE id = ?', (tx_id,))
        if tx is None or not access.can_see_transaction(tx, current_person):
            continue
        execute('DELETE FROM transactions WHERE id = ?', (tx_id,))
        deleted += 1
    return {'deleted': deleted, 'skipped': len(ids) - deleted}


def _match_account_by_iban(account_info: Optional[Dict[str, Any]]) -> Optional[int]:
    """Cerca un conto gia' censito con lo stesso IBAN individuato dall'AI
    nell'intestazione dell'estratto conto, per suggerire/selezionare il conto
    automaticamente quando l'utente non ne ha scelto uno."""
    if not account_info or not account_info.get('iban'):
        return None
    iban = str(account_info['iban']).replace(' ', '').upper()
    match = fetchone("SELECT id FROM accounts WHERE REPLACE(UPPER(iban), ' ', '') = ?", (iban,))
    return match['id'] if match else None


def _transaction_import_hash(account_id: Optional[int], date: str, amount: float, description: str) -> str:
    """Impronta deterministica di una transazione importata (stesso conto,
    data, importo arrotondato al centesimo, descrizione normalizzata): usata
    con il vincolo UNIQUE su transactions.import_hash per scartare i
    duplicati. La colonna esisteva gia' nello schema (pensata per il futuro
    sync bancario, mai implementato - vedi /api/banksync/sync) ma
    _finalize_import non la calcolava mai, quindi il vincolo UNIQUE non
    scartava nulla (piu' righe con import_hash NULL non sono in conflitto tra
    loro in SQLite) - bug reale: reimportare lo stesso estratto, o importare
    due estratti con periodi che si sovrappongono, duplicava silenziosamente
    ogni transazione. Normalizziamo la descrizione (minuscolo, spazi
    collassati) perche' la stessa causale puo' arrivare con spaziatura/maiuscole
    leggermente diverse tra un tentativo di import e l'altro (es. una ricetta
    regex diversa generata dall'AI la seconda volta)."""
    normalized_description = re.sub(r'\s+', ' ', description or '').strip().lower()
    raw = f'{account_id}|{date}|{round(float(amount), 2)}|{normalized_description}'
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


def _finalize_import(
    parsed: List[Dict[str, Any]],
    data: bytes,
    filename: str,
    content_type: Optional[str],
    account_id: Optional[int],
    import_source: str,
    used_ai: bool,
    detected_account: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Salva il file, inserisce le transazioni parsate e prova gli arricchimenti
    automatici (carta di credito collegata, ricevute email in attesa). Condivisa
    da import CSV/XLSX (sincrono) e import PDF (streaming).

    detected_account e' quanto l'AI ha individuato nell'intestazione del PDF
    (bankName/iban/cardNumber): solo informativo, non influisce sul conto gia'
    risolto in account_id (la selezione dell'account tramite IBAN avviene
    prima di chiamare questa funzione, vedi _match_account_by_iban)."""
    account = fetchone('SELECT ownership, owner_id FROM accounts WHERE id = ?', (account_id,)) if account_id else None
    # Un conto personale segrega le sue transazioni come spesa personale del suo
    # intestatario: non devono comparire tra le spese comuni. Vedi access.py.
    is_personal_account = bool(account and account['ownership'] == 'personal')
    destination = 'personal' if is_personal_account else 'family'
    paid_by_person_id = account['owner_id'] if is_personal_account else None

    batch_id = uuid.uuid4().hex
    safe_name = filename.replace('/', '_').replace('\\', '_')
    stored_path = config.DOCUMENTS_DIR / f'{batch_id}_{safe_name}'
    stored_path.write_bytes(data)
    doc_cursor = db.conn.execute(
        'INSERT INTO documents (filename, stored_path, mime_type, size_bytes, account_id, import_batch_id) VALUES (?, ?, ?, ?, ?, ?)',
        (filename, str(stored_path), content_type, len(data), account_id, batch_id),
    )
    document_id = doc_cursor.lastrowid

    # Riga di pagamento saldo carta rilevata da pdf_import (isCardSettlement,
    # vedi _CARD_SETTLEMENT_RE): e' strutturalmente un trasferimento interno
    # (non una spesa/entrata da categorizzare), quindi le assegnamo subito la
    # categoria Trasferimenti con la stessa confidenza 1.0 usata per i
    # giroconti riconosciuti per IBAN in categorize.categorize_batch, invece
    # di lasciarla passare per il matching a parole chiave.
    transfer_category = fetchone("SELECT id FROM categories WHERE is_active = 1 AND type = 'transfer' LIMIT 1")
    transfer_category_id = transfer_category['id'] if transfer_category else None

    imported = 0
    duplicates = 0
    for tx in parsed:
        try:
            is_card_settlement = bool(tx.get('isCardSettlement')) and transfer_category_id is not None
            import_hash = _transaction_import_hash(account_id, tx['date'], tx['amount'], tx['description'])
            cursor = db.conn.execute(
                'INSERT OR IGNORE INTO transactions (date, amount, description_raw, merchant_name, account_id, destination, paid_by_person_id, is_confirmed, import_source, import_batch_id, notes, document_id, ai_category_id, ai_confidence, import_hash) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (
                    tx['date'],
                    float(tx['amount']),
                    tx['description'],
                    tx['description'],
                    account_id,
                    destination,
                    paid_by_person_id,
                    0,
                    import_source,
                    batch_id,
                    None,
                    document_id,
                    transfer_category_id if is_card_settlement else None,
                    1.0 if is_card_settlement else None,
                    import_hash,
                ),
            )
            if cursor.rowcount:
                imported += 1
            else:
                duplicates += 1
        except Exception:
            continue
    db.conn.commit()

    # Se questo conto e' il conto di appoggio di una carta di credito, prova a
    # riconoscere l'addebito riepilogativo mensile per evitare la doppia conta
    # (spesa gia' registrata sulla carta + addebito unico sul c/c).
    suggested_transfers = []
    if account_id:
        linked_cards = fetchall('SELECT id, name FROM accounts WHERE settlement_account_id = ?', (account_id,))
        if linked_cards:
            batch_txs = fetchall(
                'SELECT id, date, amount, description_raw FROM transactions WHERE import_batch_id = ? AND amount < 0',
                (batch_id,),
            )
            for card in linked_cards:
                for candidate in batch_txs:
                    window_start = (parse_date(candidate['date']) - timedelta(days=45)).date().isoformat()
                    card_total = fetchone(
                        'SELECT COALESCE(SUM(ABS(amount)),0) AS total FROM transactions '
                        'WHERE account_id = ? AND amount < 0 AND date > ? AND date <= ?',
                        (card['id'], window_start, candidate['date']),
                    )['total']
                    if card_total <= 0:
                        continue
                    tolerance = max(card_total * 0.02, 3.0)
                    if abs(card_total - abs(candidate['amount'])) <= tolerance:
                        suggested_transfers.append({
                            'transactionId': candidate['id'],
                            'description': candidate['description_raw'],
                            'amount': candidate['amount'],
                            'date': candidate['date'],
                            'cardAccountId': card['id'],
                            'cardAccountName': card['name'],
                            'matchedCardTotal': round(card_total, 2),
                        })

    # Riprova ad abbinare le ricevute email (PayPal/Amazon/...) ancora in attesa
    # alle transazioni appena importate (es. l'estratto conto arriva dopo la mail).
    enriched_from_email = email_enrich.match_pending_receipts_for_batch(batch_id)

    # Suggerisce una categoria (parole chiave, poi AI solo per il resto) per le
    # transazioni appena importate: restano da confermare (is_confirmed=0), il
    # banner "categorizzate da AI" del frontend le mostra per l'approvazione.
    ai_categorized = categorize.categorize_batch(batch_id)

    return {
        'count': imported,
        'total': len(parsed),
        'duplicates': duplicates,
        'usedAi': used_ai,
        'bank': (detected_account or {}).get('bankName'),
        'signWarning': (detected_account or {}).get('signWarning'),
        'reconciliationWarning': (detected_account or {}).get('reconciliationWarning'),
        'detectedAccount': detected_account,
        'accountId': account_id,
        'filename': filename,
        'preview': parsed[:5],
        'suggestedTransfers': suggested_transfers,
        'enrichedFromEmail': enriched_from_email,
        'aiCategorized': ai_categorized,
    }


@app.post('/api/transactions/import')
def import_transactions(file: UploadFile = File(...), accountId: Optional[str] = Form(None)):
    data = file.file.read()
    name = file.filename.lower()
    used_ai = False
    detected_account = None
    account_id = ensure_int(accountId)
    if name.endswith('.pdf'):
        text = pdf_import.extract_pdf_text(data)
        if not text.strip():
            raise HTTPException(
                status_code=400,
                detail='PDF non leggibile o vuoto. Assicurati che il PDF non sia scansionato come immagine.',
            )
        try:
            detected_account, parsed = pdf_import.ai_extract_transactions_from_pdf(text, file.filename, data)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        used_ai = True
        import_source = 'pdf'
        account_id = account_id or _match_account_by_iban(detected_account)
    elif (name.endswith('.xlsx') or name.endswith('.xls')) and looks_like_meal_voucher_export(data):
        parsed = parse_rows_from_meal_voucher(data)
        import_source = 'meal_voucher'
    elif name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            parsed = parse_rows_from_xlsx(data)
            import_source = 'excel'
        except zipfile.BadZipFile:
            # Alcune banche esportano un file .xlsx che in realta' e' un testo
            # CSV rinominato (non uno zip/xlsx vero) - bug reale trovato su un
            # export reale che openpyxl rifiutava con BadZipFile.
            try:
                parsed = parse_rows_from_csv(data)
            except ValueError as e:
                raise HTTPException(status_code=400, detail=str(e))
            import_source = 'csv'
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    elif looks_like_cbi(data):
        detected_account, parsed = parse_rows_from_cbi(data)
        import_source = 'cbi'
        account_id = account_id or _match_account_by_iban(detected_account)
    else:
        try:
            parsed = parse_rows_from_csv(data)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        import_source = 'csv'
    if not parsed:
        raise HTTPException(status_code=422, detail='Nessuna transazione trovata nel file')

    return _finalize_import(
        parsed, data, file.filename, file.content_type, account_id, import_source, used_ai, detected_account,
    )


@app.post('/api/transactions/import-pdf-stream')
def import_pdf_stream(file: UploadFile = File(...), accountId: Optional[str] = Form(None)):
    """Come /api/transactions/import per i PDF, ma risponde con un flusso SSE a
    fasi (estrazione testo -> analisi formato con AI -> pattern applicato in
    locale) invece che con un'unica risposta: utile perche' la chiamata AI, pur
    piccola (analizza solo un campione per ricavare il pattern di estrazione),
    puo' comunque richiedere qualche secondo. Evento finale 'done' con lo
    stesso payload di /api/transactions/import, oppure 'error'."""
    data = file.file.read()
    filename = file.filename
    content_type = file.content_type
    account_id = ensure_int(accountId)

    def sse(event: str, payload: Dict[str, Any]) -> str:
        return f'event: {event}\ndata: {json.dumps(payload)}\n\n'

    def event_stream():
        text = pdf_import.extract_pdf_text(data)
        if not text.strip():
            yield sse('error', {'detail': 'PDF non leggibile o vuoto. Assicurati che il PDF non sia scansionato come immagine.'})
            return
        yield sse('stage', {'message': "Testo estratto dal PDF, analisi del formato con l'AI..."})

        try:
            detected_account, parsed = pdf_import.ai_extract_transactions_from_pdf(text, filename, data)
        except ValueError as e:
            yield sse('error', {'detail': str(e)})
            return

        if detected_account:
            yield sse('account', detected_account)
        yield sse('progress', {'count': len(parsed)})

        resolved_account_id = account_id or _match_account_by_iban(detected_account)
        result = _finalize_import(parsed, data, filename, content_type, resolved_account_id, 'pdf', True, detected_account)
        yield sse('done', result)

    return StreamingResponse(event_stream(), media_type='text/event-stream')


@app.get('/api/documents')
def list_documents(request: Request):
    params = request.query_params
    filters = []
    args: List[Any] = []
    if account_id := ensure_int(params.get('accountId')):
        filters.append('d.account_id = ?')
        args.append(account_id)
    if transaction_id := ensure_int(params.get('transactionId')):
        filters.append('d.transaction_id = ?')
        args.append(transaction_id)
    # period_start/period_end/tx_count vengono dalle transazioni vere legate al
    # documento (document_id), non da un periodo dichiarato nel nome del file o
    # nel PDF (che l'addon non riparsa dopo l'import): e' il dato oggettivo di
    # quali date sono state DAVVERO importate, utile per la vista di copertura
    # per conto (vedi getAccountCoverage nel frontend) che mostra da-a-a per
    # ogni documento e i buchi tra un estratto e il successivo.
    sql = (
        'SELECT d.id, d.filename, d.mime_type, d.size_bytes, d.account_id, d.import_batch_id, d.transaction_id, d.uploaded_at, '
        'MIN(t.date) AS period_start, MAX(t.date) AS period_end, COUNT(t.id) AS tx_count '
        'FROM documents d LEFT JOIN transactions t ON t.document_id = d.id'
    )
    if filters:
        sql += ' WHERE ' + ' AND '.join(filters)
    sql += ' GROUP BY d.id ORDER BY d.uploaded_at DESC'
    return fetchall(sql, tuple(args))


@app.post('/api/transactions/{transaction_id}/documents')
def upload_transaction_document(transaction_id: int, request: Request, file: UploadFile = File(...)):
    """Allega manualmente un file (es. foto di uno scontrino) a una transazione
    specifica: a differenza del documento sorgente di un import (1 documento -> N
    transazioni), qui la relazione e' 1 transazione -> N allegati."""
    tx = fetchone('SELECT * FROM transactions WHERE id = ?', (transaction_id,))
    if tx is None or not access.can_see_transaction(tx, access.get_current_person(request)):
        raise HTTPException(status_code=404, detail='Not found')
    data = file.file.read()
    safe_name = file.filename.replace('/', '_').replace('\\', '_')
    stored_path = config.DOCUMENTS_DIR / f'{uuid.uuid4().hex}_{safe_name}'
    stored_path.write_bytes(data)
    cursor = db.conn.execute(
        'INSERT INTO documents (filename, stored_path, mime_type, size_bytes, account_id, transaction_id) VALUES (?, ?, ?, ?, ?, ?)',
        (file.filename, str(stored_path), file.content_type, len(data), tx['account_id'], transaction_id),
    )
    db.conn.commit()
    return JSONResponse(status_code=201, content=fetchone('SELECT id, filename, mime_type, size_bytes, account_id, import_batch_id, transaction_id, uploaded_at FROM documents WHERE id = ?', (cursor.lastrowid,)))


@app.get('/api/documents/{document_id}/download')
def download_document(document_id: int):
    doc = fetchone('SELECT * FROM documents WHERE id = ?', (document_id,))
    if doc is None:
        raise HTTPException(status_code=404, detail='Not found')
    path = Path(doc['stored_path'])
    if not path.exists():
        raise HTTPException(status_code=404, detail='File non trovato su disco')
    return FileResponse(str(path), filename=doc['filename'], media_type=doc['mime_type'] or 'application/octet-stream')


@app.delete('/api/documents/{document_id}')
def delete_document(document_id: int):
    doc = fetchone('SELECT * FROM documents WHERE id = ?', (document_id,))
    if doc is None:
        raise HTTPException(status_code=404, detail='Not found')
    path = Path(doc['stored_path'])
    if path.exists():
        path.unlink()
    execute('DELETE FROM documents WHERE id = ?', (document_id,))
    return JSONResponse(status_code=204, content=None)


# Frammento SQL riusato in tutti i report/aggregati spese-entrate: esclude i
# 'transfer' (gia' cosi' prima) e i checkpoint 'opening_balance' (saldo
# iniziale annuale, vedi _compute_account_balances) - ne' gli uni ne' gli
# altri sono spese/entrate reali. Sicuro anche quando la join su categories e'
# una LEFT JOIN (c.type puo' essere NULL) o una JOIN normale (mai NULL, il
# ramo IS NULL resta innocuo).
_NON_SPEND_TYPES_SQL = "(c.type IS NULL OR c.type NOT IN ('transfer', 'opening_balance'))"


def _period_expense_income(pattern: str, vis_clause_t: str, vis_args_t: tuple, account_clause: str = '', account_args: tuple = ()) -> Dict[str, float]:
    """Totali spese/entrate per un periodo (usato per il confronto mese/anno
    precedente): stessa logica di calcolo di report_summary (esclude i
    'transfer' e rispetta la visibilita' personale/famiglia), ma senza il
    dettaglio per categoria - qui serve solo il totale."""
    row = fetchone(
        'SELECT '
        'COALESCE(SUM(CASE WHEN t.amount<0 THEN ABS(t.amount) ELSE 0 END),0) AS total_expenses, '
        'COALESCE(SUM(CASE WHEN t.amount>0 THEN t.amount ELSE 0 END),0) AS total_income '
        'FROM transactions t LEFT JOIN categories c ON c.id = t.category_id '
        f"WHERE t.date LIKE ? AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL} AND {vis_clause_t}{account_clause}",
        (pattern,) + vis_args_t + account_args,
    )
    return {'total_expenses': row['total_expenses'], 'total_income': row['total_income']}


@app.get('/api/reports/summary')
def report_summary(request: Request):
    month = request.query_params.get('month') or datetime.utcnow().strftime('%Y-%m')
    pattern = f'{month}%'
    year_pattern = f'{month[:4]}%'
    current_person = access.get_current_person(request)
    vis_clause_t, vis_args_t = access.transaction_visibility(current_person, alias='t')
    account_id = ensure_int(request.query_params.get('accountId'))
    account_clause = ' AND t.account_id = ?' if account_id else ''
    account_args = (account_id,) if account_id else ()
    # Mese precedente: sottrarre un giorno dal primo del mese corrente da'
    # sempre l'ultimo giorno del mese precedente, gestendo correttamente anche
    # il cambio di anno (es. gennaio -> dicembre dell'anno prima) senza dover
    # calcolare a mano i giorni per ogni mese.
    year_n, month_n = int(month[:4]), int(month[5:7])
    previous_month_str = (datetime(year_n, month_n, 1) - timedelta(days=1)).strftime('%Y-%m')
    previous_year_month_str = f'{year_n - 1}-{month[5:7]}'
    previous_month = _period_expense_income(f'{previous_month_str}%', vis_clause_t, vis_args_t, account_clause, account_args)
    previous_year_same_month = _period_expense_income(f'{previous_year_month_str}%', vis_clause_t, vis_args_t, account_clause, account_args)
    # Le transazioni categoria 'transfer' (es. pagamento carta di credito verso il
    # conto di appoggio) non sono spese reali: la spesa e' gia' contata sulla carta.
    totals = fetchone(
        'SELECT '
        'COALESCE(SUM(CASE WHEN t.date LIKE ? AND t.amount<0 THEN ABS(t.amount) ELSE 0 END),0) AS total_expenses, '
        'COALESCE(SUM(CASE WHEN t.date LIKE ? AND t.amount>0 THEN t.amount ELSE 0 END),0) AS total_income, '
        'COALESCE(SUM(CASE WHEN t.amount<0 THEN ABS(t.amount) ELSE 0 END),0) AS total_expenses_year, '
        'COALESCE(SUM(CASE WHEN t.amount>0 THEN t.amount ELSE 0 END),0) AS total_income_year '
        'FROM transactions t LEFT JOIN categories c ON c.id = t.category_id '
        f"WHERE t.date LIKE ? AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL} AND {vis_clause_t}{account_clause}",
        (pattern, pattern, year_pattern) + vis_args_t + account_args,
    )
    by_category = fetchall(
        'SELECT c.id, c.name, c.icon, c.color, c.budget_monthly, c.budget_annual, '
        'COALESCE(SUM(CASE WHEN t.date LIKE ? THEN ABS(t.amount) ELSE 0 END),0) AS spent, '
        'COALESCE(SUM(ABS(t.amount)),0) AS spent_year '
        'FROM transactions t JOIN categories c ON c.id = t.category_id '
        f"WHERE t.date LIKE ? AND t.amount<0 AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL} AND {vis_clause_t}{account_clause} "
        'GROUP BY c.id ORDER BY spent DESC',
        (pattern, year_pattern) + vis_args_t + account_args,
    )
    by_destination = fetchall(
        'SELECT t.destination, COALESCE(SUM(ABS(t.amount)),0) AS total FROM transactions t '
        'LEFT JOIN categories c ON c.id = t.category_id '
        f"WHERE t.date LIKE ? AND t.amount<0 AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL} AND {vis_clause_t}{account_clause} "
        'GROUP BY t.destination',
        (pattern,) + vis_args_t + account_args,
    )
    return {
        'month': month,
        'total_expenses': totals['total_expenses'],
        'total_income': totals['total_income'],
        'total_expenses_year': totals['total_expenses_year'],
        'total_income_year': totals['total_income_year'],
        'byCategory': by_category,
        'byDestination': by_destination,
        'previousMonth': previous_month,
        'previousYearSameMonth': previous_year_same_month,
    }


@app.get('/api/reports/top-merchants')
def report_top_merchants(request: Request):
    month = request.query_params.get('month') or datetime.utcnow().strftime('%Y-%m')
    pattern = f'{month}%'
    limit = ensure_int(request.query_params.get('limit')) or 10
    vis_clause_t, vis_args_t = access.transaction_visibility(access.get_current_person(request), alias='t')
    account_id = ensure_int(request.query_params.get('accountId'))
    account_clause = ' AND t.account_id = ?' if account_id else ''
    account_args = (account_id,) if account_id else ()
    rows = fetchall(
        'SELECT t.merchant_name, COUNT(*) AS occurrences, COALESCE(SUM(ABS(t.amount)),0) AS total '
        'FROM transactions t LEFT JOIN categories c ON c.id = t.category_id '
        f"WHERE t.date LIKE ? AND t.amount<0 AND t.is_confirmed=1 AND t.merchant_name IS NOT NULL "
        f"AND {_NON_SPEND_TYPES_SQL} AND {vis_clause_t}{account_clause} "
        'GROUP BY t.merchant_name ORDER BY total DESC LIMIT ?',
        (pattern,) + vis_args_t + account_args + (limit,),
    )
    return rows


@app.get('/api/reports/trend')
def report_trend(request: Request):
    months = ensure_int(request.query_params.get('months')) or 6
    vis_clause_t, vis_args_t = access.transaction_visibility(access.get_current_person(request), alias='t')
    account_id = ensure_int(request.query_params.get('accountId'))
    account_clause = ' AND t.account_id = ?' if account_id else ''
    account_args = (account_id,) if account_id else ()
    rows = fetchall(
        "SELECT strftime('%Y-%m', t.date) AS month, "
        "COALESCE(SUM(CASE WHEN t.amount<0 AND t.destination='family' THEN ABS(t.amount) ELSE 0 END),0) AS family, "
        "COALESCE(SUM(CASE WHEN t.amount<0 AND t.destination!='family' THEN ABS(t.amount) ELSE 0 END),0) AS personal, "
        "COALESCE(SUM(CASE WHEN t.amount>0 THEN t.amount ELSE 0 END),0) AS income "
        "FROM transactions t LEFT JOIN categories c ON c.id = t.category_id "
        f"WHERE t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL} AND {vis_clause_t}{account_clause} "
        "GROUP BY strftime('%Y-%m', t.date) ORDER BY month DESC LIMIT ?",
        vis_args_t + account_args + (months,),
    )
    return list(reversed(rows))


@app.get('/api/reports/pivot')
def report_pivot(request: Request):
    """Pivot categoria x mese per la tabella pivot in Report: righe = categorie
    di spesa, colonne = ultimi N mesi, celle = totale speso in quel mese per
    quella categoria. Endpoint separato da report_trend (che aggrega per
    destinazione family/personal, non per categoria)."""
    months_n = ensure_int(request.query_params.get('months')) or 6
    vis_clause_t, vis_args_t = access.transaction_visibility(access.get_current_person(request), alias='t')
    account_id = ensure_int(request.query_params.get('accountId'))
    account_clause = ' AND t.account_id = ?' if account_id else ''
    account_args = (account_id,) if account_id else ()

    # Elenco fisso degli ultimi N mesi (piu' vecchio per primo): serve per
    # avere colonne stabili anche per i mesi senza spese in una data
    # categoria, che la GROUP BY sotto altrimenti ometterebbe del tutto.
    today = datetime.utcnow()
    month_list = []
    y, m = today.year, today.month
    for _ in range(months_n):
        month_list.append(f'{y:04d}-{m:02d}')
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    month_list.reverse()
    earliest = f'{month_list[0]}-01'

    rows = fetchall(
        "SELECT c.id, c.name, c.icon, c.color, strftime('%Y-%m', t.date) AS month, "
        'COALESCE(SUM(ABS(t.amount)),0) AS total '
        'FROM transactions t JOIN categories c ON c.id = t.category_id '
        f"WHERE t.amount<0 AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL} AND t.date >= ? AND {vis_clause_t}{account_clause} "
        "GROUP BY c.id, month",
        (earliest,) + vis_args_t + account_args,
    )

    by_category: Dict[int, Dict[str, Any]] = {}
    for r in rows:
        cat = by_category.setdefault(r['id'], {
            'id': r['id'], 'name': r['name'], 'icon': r['icon'], 'color': r['color'],
            'totals': {mo: 0.0 for mo in month_list}, 'total': 0.0,
        })
        cat['totals'][r['month']] = r['total']
        cat['total'] += r['total']

    result_rows = sorted(by_category.values(), key=lambda c: c['total'], reverse=True)
    return {'months': month_list, 'rows': result_rows}


@app.get('/api/reports/balance')
def report_balance(request: Request):
    month = request.query_params.get('month') or datetime.utcnow().strftime('%Y-%m')
    period = request.query_params.get('period') or 'month'
    pattern = '%' if period == 'all' else f'{month}%'
    persons = fetchall('SELECT id, name, color FROM persons ORDER BY id')
    if not persons:
        return {'month': month, 'period': period, 'persons': [], 'debt': None}
    tx_rows = fetchall(
        'SELECT t.paid_by_person_id, t.split_person_id, t.amount, t.destination, t.split_ratio, '
        'c.type AS cat_type, a.ownership '
        'FROM transactions t LEFT JOIN categories c ON c.id = t.category_id LEFT JOIN accounts a ON a.id = t.account_id '
        'WHERE t.date LIKE ? AND t.is_confirmed=1',
        (pattern,),
    )
    stats = {p['id']: {'id': p['id'], 'name': p['name'], 'color': p['color'], 'contributed': 0.0, 'personalSpent': 0.0, 'familySpent': 0.0} for p in persons}
    num_persons = max(len(persons), 2)
    for tx in tx_rows:
        pid = tx['paid_by_person_id']
        amount = tx['amount'] or 0
        dest = tx['destination']
        cat_type = tx['cat_type']
        ownership = tx['ownership']
        if cat_type == 'transfer':
            continue
        if amount > 0:
            if ownership == 'shared' and pid and pid in stats:
                stats[pid]['contributed'] += amount
        else:
            abs_amount = abs(amount)
            if dest == 'personal':
                # Una spesa personale pesa sul bilancio comune solo se pagata
                # con soldi del conto condiviso (ha eroso il fondo comune).
                # Se pagata da un conto personale, sono soldi gia' suoi: non
                # deve incidere su quanto deve/e' dovuto tra le persone.
                if ownership == 'shared' and pid and pid in stats:
                    stats[pid]['personalSpent'] += abs_amount
            elif dest == 'split' and tx['split_person_id'] and tx['split_person_id'] in stats:
                # Spesa divisa solo tra chi ha pagato e la persona indicata,
                # secondo split_ratio (quota di chi ha pagato), non tra tutti.
                # Chi ha pagato ha anticipato l'intera cifra: va accreditato per
                # il totale (come un versamento), non solo per la propria quota,
                # altrimenti risulterebbe lui il debitore invece del creditore.
                ratio = tx['split_ratio'] if tx['split_ratio'] is not None else 0.5
                other_pid = tx['split_person_id']
                if pid and pid in stats:
                    stats[pid]['contributed'] += abs_amount
                    stats[pid]['familySpent'] += abs_amount * ratio
                stats[other_pid]['familySpent'] += abs_amount * (1 - ratio)
            else:
                share = abs_amount / num_persons
                for p in persons:
                    stats[p['id']]['familySpent'] += share
    result = []
    for s in stats.values():
        net = s['contributed'] - s['personalSpent'] - s['familySpent']
        result.append({
            **s,
            'contributed': round(s['contributed'], 2),
            'personalSpent': round(s['personalSpent'], 2),
            'familySpent': round(s['familySpent'], 2),
            'net': round(net, 2),
        })
    debt = None
    if len(result) == 2:
        a, b = result[0], result[1]
        diff = round((a['net'] - b['net']), 2)
        if abs(diff) > 0.01:
            debtor, creditor = (a, b) if diff < 0 else (b, a)
            debt = {'debtor': debtor['name'], 'creditor': creditor['name'], 'amount': round(abs(diff) / 2, 2)}
    # Il calcolo del debito sopra usa il dettaglio (contributed/personalSpent/
    # familySpent) di TUTTE le persone - e' necessario per essere corretto (le
    # spese personali pagate dal conto comune incidono sul fondo comune anche
    # se non sono "tue"). Ma quel dettaglio riga per riga di un'altra persona
    # e' esattamente cio' che non deve essere visibile (bug reale segnalato
    # dall'utente): nella risposta esponiamo solo la card della persona
    # corrente, il banner "chi deve cosa" (gia' senza dettagli) resta invariato
    # per tutti.
    current_person = access.get_current_person(request)
    visible_persons = [p for p in result if p['id'] == current_person['id']] if current_person else result
    return {'month': month, 'period': period, 'persons': visible_persons, 'debt': debt}


@app.get('/api/reports/subscriptions')
def report_subscriptions(request: Request):
    vis_clause_t, vis_args_t = access.transaction_visibility(access.get_current_person(request), alias='t')
    rows = fetchall(
        'SELECT t.merchant_name, ABS(t.amount) AS amount, COUNT(*) AS occurrences, MAX(t.date) AS last_date '
        'FROM transactions t LEFT JOIN categories c ON c.id = t.category_id '
        f"WHERE t.amount<0 AND t.is_confirmed=1 AND t.merchant_name IS NOT NULL "
        f"AND {_NON_SPEND_TYPES_SQL} AND {vis_clause_t} "
        'GROUP BY t.merchant_name, ROUND(ABS(t.amount),0) HAVING occurrences>=2 ORDER BY amount DESC',
        vis_args_t,
    )
    total_monthly = sum(row['amount'] for row in rows)
    return {'subscriptions': rows, 'totalMonthly': total_monthly}


@app.post('/api/reports/balance/{month}/settle')
def report_balance_settle(month: str):
    return {'settled': True, 'month': month}


# Dimensioni disponibili nel report builder: 'select' e' anche l'espressione di
# GROUP BY (deve restare l'espressione grezza, non l'alias, perche' SQLite non
# accetta sempre un alias di SELECT dentro GROUP BY in presenza di funzioni
# come strftime). 'join' e' vuoto per le dimensioni gia' su transactions.
_REPORT_DIMENSIONS = {
    'category': {
        'select': "COALESCE(categories.name, 'Senza categoria')",
        'join': 'LEFT JOIN categories ON categories.id = transactions.category_id',
    },
    'account': {
        'select': 'accounts.name',
        'join': 'JOIN accounts ON accounts.id = transactions.account_id',
    },
    'person': {
        'select': "COALESCE(persons.name, 'Non specificato')",
        'join': 'LEFT JOIN persons ON persons.id = transactions.paid_by_person_id',
    },
    'destination': {
        'select': 'transactions.destination',
        'join': '',
    },
    'month': {
        'select': "strftime('%Y-%m', transactions.date)",
        'join': '',
    },
    'day': {
        'select': 'transactions.date',
        'join': '',
    },
    'merchant': {
        'select': "COALESCE(transactions.merchant_name, transactions.description_raw, 'Sconosciuto')",
        'join': '',
    },
}
_REPORT_METRICS = {'sum', 'count', 'avg'}


@app.post('/api/reports/query')
def report_query(payload: Dict[str, Any], request: Request):
    """Query generica per il report builder: dimensioni/filtri/metrica arrivano
    dal frontend, ma solo come CHIAVI whitelisted in _REPORT_DIMENSIONS/
    _REPORT_METRICS - i valori concreti (date, id, destination) restano sempre
    parametri bind, mai concatenati nella query. Max 2 dimensioni: oltre non
    aggiunge leggibilita' a una tabella/grafico pensato per essere letto a
    colpo d'occhio."""
    dimensions = [d for d in (payload.get('dimensions') or []) if d in _REPORT_DIMENSIONS][:2]
    if not dimensions:
        raise HTTPException(status_code=400, detail='Serve almeno una dimensione valida')
    metric = payload.get('metric') if payload.get('metric') in _REPORT_METRICS else 'sum'
    absolute = bool(payload.get('absolute'))
    filters_in = payload.get('filters') or {}

    joins = []
    for dim in dimensions:
        join = _REPORT_DIMENSIONS[dim]['join']
        if join and join not in joins:
            joins.append(join)

    where = []
    args: List[Any] = []
    if date_from := filters_in.get('dateFrom'):
        where.append('transactions.date >= ?')
        args.append(date_from)
    if date_to := filters_in.get('dateTo'):
        where.append('transactions.date <= ?')
        args.append(date_to)
    if account_id := ensure_int(filters_in.get('accountId')):
        where.append('transactions.account_id = ?')
        args.append(account_id)
    if category_id := ensure_int(filters_in.get('categoryId')):
        where.append('transactions.category_id = ?')
        args.append(category_id)
    if person_id := ensure_int(filters_in.get('personId')):
        where.append('transactions.paid_by_person_id = ?')
        args.append(person_id)
    if destination := filters_in.get('destination'):
        where.append('transactions.destination = ?')
        args.append(destination)
    tx_type = filters_in.get('type')
    if tx_type == 'expense':
        where.append('transactions.amount < 0')
    elif tx_type == 'income':
        where.append('transactions.amount > 0')
    if filters_in.get('confirmedOnly'):
        where.append('transactions.is_confirmed = 1')

    vis_clause, vis_args = access.transaction_visibility(access.get_current_person(request), alias='transactions')
    where.append(vis_clause)
    args.extend(vis_args)

    select_exprs = [f"{_REPORT_DIMENSIONS[d]['select']} AS dim{i}" for i, d in enumerate(dimensions)]
    group_exprs = [_REPORT_DIMENSIONS[d]['select'] for d in dimensions]
    amount_expr = 'ABS(transactions.amount)' if absolute else 'transactions.amount'
    metric_expr = {
        'sum': f'COALESCE(SUM({amount_expr}),0)',
        'count': 'COUNT(*)',
        'avg': f'COALESCE(AVG({amount_expr}),0)',
    }[metric]

    sql = f"SELECT {', '.join(select_exprs)}, {metric_expr} AS value FROM transactions " + ' '.join(joins)
    if where:
        sql += ' WHERE ' + ' AND '.join(where)
    sql += f" GROUP BY {', '.join(group_exprs)} ORDER BY value DESC"

    rows = fetchall(sql, tuple(args))
    return [
        {**{f'dim{i}': row[f'dim{i}'] for i in range(len(dimensions))}, 'value': row['value']}
        for row in rows
    ]


@app.get('/api/reports/custom')
def list_saved_reports():
    rows = fetchall('SELECT * FROM saved_reports ORDER BY updated_at DESC')
    for row in rows:
        row['config'] = json.loads(row.pop('config_json'))
    return rows


@app.post('/api/reports/custom')
def create_saved_report(payload: Dict[str, Any]):
    name = (payload.get('name') or '').strip()
    if not name:
        raise HTTPException(status_code=400, detail='Nome obbligatorio')
    cursor = db.conn.execute(
        'INSERT INTO saved_reports (name, config_json) VALUES (?, ?)',
        (name, json.dumps(payload.get('config') or {})),
    )
    db.conn.commit()
    row = fetchone('SELECT * FROM saved_reports WHERE id = ?', (cursor.lastrowid,))
    row['config'] = json.loads(row.pop('config_json'))
    return JSONResponse(status_code=201, content=row)


@app.put('/api/reports/custom/{report_id}')
def update_saved_report(report_id: int, payload: Dict[str, Any]):
    existing = fetchone('SELECT * FROM saved_reports WHERE id = ?', (report_id,))
    if existing is None:
        raise HTTPException(status_code=404, detail='Not found')
    name = (payload.get('name') or existing['name']).strip()
    config = payload.get('config') if 'config' in payload else json.loads(existing['config_json'])
    execute(
        "UPDATE saved_reports SET name = ?, config_json = ?, updated_at = (datetime('now')) WHERE id = ?",
        (name, json.dumps(config), report_id),
    )
    row = fetchone('SELECT * FROM saved_reports WHERE id = ?', (report_id,))
    row['config'] = json.loads(row.pop('config_json'))
    return row


@app.delete('/api/reports/custom/{report_id}')
def delete_saved_report(report_id: int):
    execute('DELETE FROM saved_reports WHERE id = ?', (report_id,))
    return JSONResponse(status_code=204, content=None)


def get_nordigen_credentials():
    """Le credenziali GoCardless/Nordigen inserite in Impostazioni (tabella
    settings) hanno la precedenza su quelle statiche da config.yaml/opzioni
    addon, cosi' l'utente puo' configurarle dall'app senza toccare HA."""
    setting = fetchone("SELECT value FROM settings WHERE key = 'nordigen_secret_id'")
    secret_id = json.loads(setting['value']) if setting else None
    setting = fetchone("SELECT value FROM settings WHERE key = 'nordigen_secret_key'")
    secret_key = json.loads(setting['value']) if setting else None
    return (secret_id or config.NORDIGEN_SECRET_ID) or None, (secret_key or config.NORDIGEN_SECRET_KEY) or None


@app.get('/api/banksync/status')
def banksync_status():
    connected = fetchone('SELECT COUNT(*) AS count FROM accounts WHERE nordigen_id IS NOT NULL AND is_active = 1')['count']
    last_sync = fetchone('SELECT * FROM bank_sync_log ORDER BY synced_at DESC LIMIT 1')
    secret_id, secret_key = get_nordigen_credentials()
    return {
        'connected': connected,
        'hasCredentials': bool(secret_id and secret_key),
        'lastSync': last_sync['synced_at'] if last_sync else None,
        'lastError': last_sync['error'] if last_sync else None,
    }


def get_nordigen_token():
    secret_id, secret_key = get_nordigen_credentials()
    if not secret_id or not secret_key:
        raise HTTPException(status_code=400, detail='Nordigen non configurato. Aggiungi le credenziali nel Setup.')
    response = httpx.post(
        'https://bankaccountdata.gocardless.com/api/v2/token/new/',
        json={'secret_id': secret_id, 'secret_key': secret_key},
        timeout=30.0,
    )
    if response.status_code != 200:
        raise HTTPException(status_code=500, detail='Nordigen auth failed')
    return response.json()['access']


@app.get('/api/banksync/banks')
def banksync_banks(country: str = 'IT'):
    token = get_nordigen_token()
    response = httpx.get(
        f'https://bankaccountdata.gocardless.com/api/v2/institutions/?country={country}',
        headers={'Authorization': f'Bearer {token}'},
        timeout=30.0,
    )
    return response.json()


@app.post('/api/banksync/connect')
def banksync_connect():
    raise HTTPException(status_code=501, detail='Bank connect implementato in Fase 5')


@app.get('/api/banksync/callback')
def banksync_callback():
    raise HTTPException(status_code=501, detail='OAuth callback implementato in Fase 5')


@app.post('/api/banksync/sync')
def banksync_sync():
    raise HTTPException(status_code=501, detail='Sync implementato in Fase 5')


@app.get('/api/banksync/log')
def banksync_log(limit: int = 50):
    return fetchall('SELECT * FROM bank_sync_log ORDER BY synced_at DESC LIMIT ?', (limit,))


@app.get('/api/ha/whoami')
def ha_whoami(request: Request):
    ha_user_id = request.headers.get('x-remote-user-id')
    ha_user_name = request.headers.get('x-remote-user-name')
    ha_display_name = request.headers.get('x-remote-user-display-name')
    matched = fetchone('SELECT * FROM persons WHERE ha_user_id = ?', (ha_user_id,)) if ha_user_id else None
    return {
        'haUserId': ha_user_id,
        'haUserName': ha_user_name,
        'haUserDisplayName': ha_display_name,
        'matchedPersonId': matched['id'] if matched else None,
    }


@app.get('/api/ha/sensors')
def ha_sensors():
    month = datetime.utcnow().strftime('%Y-%m')
    today = datetime.utcnow().strftime('%Y-%m-%d')
    expenses = fetchone(
        "SELECT COALESCE(SUM(ABS(t.amount)),0) AS total FROM transactions t LEFT JOIN categories c ON c.id = t.category_id "
        f"WHERE t.date LIKE ? AND t.amount<0 AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL}",
        (f'{month}%',),
    )['total']
    today_total = fetchone(
        "SELECT COALESCE(SUM(ABS(t.amount)),0) AS total FROM transactions t LEFT JOIN categories c ON c.id = t.category_id "
        f"WHERE t.date = ? AND t.amount<0 AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL}",
        (today,),
    )['total']
    shared_account_ids = {row['id'] for row in fetchall(
        'SELECT id FROM accounts WHERE ownership = ? AND is_active = 1', ('shared',),
    )}
    account_balances = _compute_account_balances()
    balance = sum(v for k, v in account_balances.items() if k in shared_account_ids)
    pending = fetchone('SELECT COUNT(*) AS count FROM transactions WHERE is_confirmed = 0')['count']
    year_pattern = f'{month[:4]}%'
    expenses_year = fetchone(
        "SELECT COALESCE(SUM(ABS(t.amount)),0) AS total FROM transactions t LEFT JOIN categories c ON c.id = t.category_id "
        f"WHERE t.date LIKE ? AND t.amount<0 AND t.is_confirmed=1 AND {_NON_SPEND_TYPES_SQL}",
        (year_pattern,),
    )['total']
    over_budget = fetchall(
        'SELECT c.name FROM categories c WHERE c.budget_monthly IS NOT NULL AND c.is_active = 1 AND (SELECT COALESCE(SUM(ABS(t.amount)),0) FROM transactions t WHERE t.category_id=c.id AND t.date LIKE ? AND t.amount<0) > c.budget_monthly',
        (f'{month}%',),
    )
    over_budget_annual = fetchall(
        'SELECT c.name FROM categories c WHERE c.budget_annual IS NOT NULL AND c.is_active = 1 AND (SELECT COALESCE(SUM(ABS(t.amount)),0) FROM transactions t WHERE t.category_id=c.id AND t.date LIKE ? AND t.amount<0) > c.budget_annual',
        (year_pattern,),
    )
    return {
        'spese_mese': round(expenses, 2),
        'spese_oggi': round(today_total, 2),
        'spese_anno': round(expenses_year, 2),
        'saldo_comuni': round(balance, 2),
        'pending_review': pending,
        'budget_ok': len(over_budget) == 0,
        'over_budget': [row['name'] for row in over_budget],
        'budget_ok_annual': len(over_budget_annual) == 0,
        'over_budget_annual': [row['name'] for row in over_budget_annual],
        'sync_ok': True,
        'month': month,
    }


@app.post('/api/ha/sync-persons')
def ha_sync_persons():
    if not config.SUPERVISOR_TOKEN:
        raise HTTPException(status_code=503, detail='SUPERVISOR_TOKEN non disponibile — funzione attiva solo dentro Home Assistant')
    try:
        response = httpx.get(
            'http://supervisor/core/api/states',
            headers={'Authorization': f'Bearer {config.SUPERVISOR_TOKEN}', 'Content-Type': 'application/json'},
            timeout=30.0,
        )
    except httpx.HTTPError as e:
        raise HTTPException(
            status_code=502,
            detail=f"Impossibile raggiungere l'API di Home Assistant ({e}). Verifica che l'addon abbia il permesso "
            "'homeassistant_api: true' in config.yaml e riavvialo.",
        )
    if response.status_code == 403:
        raise HTTPException(
            status_code=502,
            detail="Accesso negato dall'API di Home Assistant. Aggiungi 'homeassistant_api: true' in config.yaml "
            "e riavvia l'addon.",
        )
    if response.status_code != 200:
        raise HTTPException(status_code=502, detail=f'HA API error ({response.status_code})')
    states = response.json()
    persons = [s for s in states if s.get('entity_id', '').startswith('person.')]
    existing = fetchall('SELECT name FROM persons')
    existing_names = {item['name'].lower() for item in existing}
    imported = 0
    for entity in persons:
        name = entity.get('attributes', {}).get('friendly_name') or entity.get('entity_id', '').split('.')[-1]
        if name.lower() in existing_names:
            continue
        db.conn.execute('INSERT INTO persons (name, email, color, is_primary) VALUES (?, ?, ?, ?)', (name, None, '#1D3557', 0))
        imported += 1
    db.conn.commit()
    return {'imported': imported, 'total': len(persons)}


@app.post('/api/ha/webhook')
def ha_webhook(payload: Dict[str, Any]):
    # Automazione HA (integrazione IMAP) che inoltra mail di conferma
    # acquisto/pagamento (PayPal, Amazon, ...) da arricchire via AI.
    if payload.get('sender') and (payload.get('subject') or payload.get('body')):
        try:
            result = email_enrich.process_incoming_email(
                payload.get('sender', ''), payload.get('subject', ''), payload.get('body', '')
            )
            return {'received': True, 'emailReceipt': result}
        except ValueError as e:
            return JSONResponse(status_code=422, content={'received': True, 'error': str(e)})
    print('HA webhook', payload)
    return {'received': True}


@app.get('/api/email-receipts')
def list_email_receipts(request: Request):
    """Elenca le ricevute email. Filtri opzionali: 'id' (una ricevuta precisa,
    usato per "vai alla mail" da una transazione anche se e' oltre il limite
    delle ultime 100) e 'transactionId' (la ricevuta abbinata a quella
    transazione, se esiste). La visibilita' segue quella della transazione
    abbinata (una ricevuta legata a una spesa personale altrui non deve
    comparire, stessa regola di access.transaction_visibility usata ovunque
    per le transazioni) - una ricevuta senza transazione abbinata (LEFT JOIN
    NULL) resta sempre visibile, non c'e' nulla di personale da nascondere."""
    params = request.query_params
    vis_clause, vis_args = access.transaction_visibility(access.get_current_person(request), alias='t')
    filters = [vis_clause]
    args: List[Any] = list(vis_args)
    if receipt_id := ensure_int(params.get('id')):
        filters.append('e.id = ?')
        args.append(receipt_id)
    if tx_id := ensure_int(params.get('transactionId')):
        filters.append('e.matched_transaction_id = ?')
        args.append(tx_id)
    if not receipt_id and not tx_id:
        # Nell'elenco generale interessano solo le mail da cui l'AI e' riuscita
        # a estrarre un importo (le altre sono rumore, es. notifiche di
        # spedizione senza ricevuta di pagamento) - una ricerca puntuale per id
        # o per transazione invece va sempre trovata, anche senza importo.
        filters.append('e.amount IS NOT NULL')
    sql = (
        'SELECT e.* FROM email_receipts e LEFT JOIN transactions t ON t.id = e.matched_transaction_id '
        'WHERE ' + ' AND '.join(filters) + ' ORDER BY e.received_at DESC LIMIT 100'
    )
    return fetchall(sql, tuple(args))


@app.post('/api/email-receipts/rematch')
def rematch_email_receipts():
    """Bottone "Riabbina mail": ritenta a mano l'abbinamento di tutte le
    ricevute email ancora in attesa, non solo quelle di un batch di import
    appena fatto (vedi email_enrich.rematch_all_pending_receipts)."""
    return {'matched': email_enrich.rematch_all_pending_receipts()}


@app.post('/api/email-receipts/{receipt_id}/unmatch')
def unmatch_email_receipt(receipt_id: int):
    """Slega una ricevuta da un abbinamento sbagliato (es. mail vecchia
    abbinata a una transazione recente per una data indovinata male dall'AI -
    vedi email_enrich.process_incoming_email). Riporta anche la transazione a
    merchant_enriched=0, altrimenti resterebbe esclusa per sempre dal pool di
    rematch_all_pending_receipts/_match_receipts_against e non potrebbe mai
    essere riabbinata alla ricevuta giusta."""
    receipt = fetchone('SELECT * FROM email_receipts WHERE id = ?', (receipt_id,))
    if receipt is None:
        raise HTTPException(status_code=404, detail='Not found')
    tx_id = receipt['matched_transaction_id']
    execute('UPDATE email_receipts SET matched_transaction_id = NULL WHERE id = ?', (receipt_id,))
    if tx_id:
        execute('UPDATE transactions SET merchant_enriched = 0 WHERE id = ?', (tx_id,))
    return {'unmatched': True}


# SPA fallback: montato DOPO tutte le route API così non le intercetta
if PUBLIC_DIR.exists():
    app.mount('/', StaticFiles(directory=str(PUBLIC_DIR)), name='public')
