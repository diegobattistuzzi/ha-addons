import io
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

from . import db

# Ordine di import: prima le tabelle referenziate (persone, categorie), poi
# quelle che le referenziano (conti, budget, transazioni) - SQLite qui non
# applica i vincoli FK, ma l'ordine resta piu' pulito e prevedibile.
SHEET_SPECS = {
    'persons': ['id', 'name', 'email', 'color', 'is_primary', 'ha_user_id'],
    'categories': [
        'id', 'code', 'name', 'icon', 'color', 'parent_id', 'budget_monthly', 'type',
        'ai_keywords', 'sort_order', 'is_active',
    ],
    'accounts': [
        'id', 'name', 'bank', 'type', 'ownership', 'owner_id', 'co_owners', 'iban', 'color',
        'nordigen_id', 'balance', 'is_active', 'settlement_account_id',
    ],
    'budgets': ['id', 'category_id', 'year_month', 'amount'],
    'transactions': [
        'id', 'date', 'amount', 'currency', 'description_raw', 'merchant_name', 'merchant_category_code',
        'category_id', 'account_id', 'destination', 'paid_by_person_id', 'split_person_id', 'split_ratio',
        'space_name', 'is_cash', 'ai_category_id', 'ai_confidence', 'is_confirmed', 'import_hash',
        'import_source', 'import_batch_id', 'reimbursement_of', 'notes', 'merchant_enriched',
        'is_reimbursable', 'reimbursed_at', 'document_id',
    ],
}
IMPORT_ORDER = ['persons', 'categories', 'accounts', 'budgets', 'transactions']


def _fetchall(query: str) -> List[Dict[str, Any]]:
    cursor = db.conn.execute(query)
    return [{k: row[k] for k in row.keys()} for row in cursor.fetchall()]


def build_backup_workbook() -> Workbook:
    """Un file .xlsx con un foglio per tabella: persone, conti, categorie,
    transazioni, budget. Le credenziali IMAP delle persone NON sono incluse."""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, columns in SHEET_SPECS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        for row in _fetchall(f'SELECT {", ".join(columns)} FROM {sheet_name} ORDER BY id'):
            ws.append([row.get(c) for c in columns])
    return wb


def parse_backup_workbook(data: bytes) -> Dict[str, List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    result: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            result[sheet_name] = []
            continue
        headers = [str(h).strip() if h is not None else '' for h in headers]
        sheet_rows = []
        for raw_row in rows_iter:
            if raw_row is None or all(v is None for v in raw_row):
                continue
            sheet_rows.append({headers[i]: raw_row[i] for i in range(len(headers)) if i < len(raw_row)})
        result[sheet_name] = sheet_rows
    return result


def _upsert_rows(table: str, columns: List[str], rows: List[Dict[str, Any]]) -> Dict[str, int]:
    inserted = updated = skipped = 0
    for row in rows:
        row_id = row.get('id')
        try:
            if row_id not in (None, ''):
                existing = db.conn.execute(f'SELECT id FROM {table} WHERE id = ?', (row_id,)).fetchone()
                if existing:
                    other_cols = [c for c in columns if c != 'id']
                    set_clause = ', '.join(f'{c} = ?' for c in other_cols)
                    values = [row.get(c) for c in other_cols]
                    db.conn.execute(f'UPDATE {table} SET {set_clause} WHERE id = ?', (*values, row_id))
                    updated += 1
                    continue
                col_list = ', '.join(columns)
                placeholders = ', '.join('?' for _ in columns)
                values = [row.get(c) for c in columns]
                db.conn.execute(f'INSERT INTO {table} ({col_list}) VALUES ({placeholders})', values)
                inserted += 1
                continue
            other_cols = [c for c in columns if c != 'id']
            col_list = ', '.join(other_cols)
            placeholders = ', '.join('?' for _ in other_cols)
            values = [row.get(c) for c in other_cols]
            db.conn.execute(f'INSERT INTO {table} ({col_list}) VALUES ({placeholders})', values)
            inserted += 1
        except Exception:
            skipped += 1
    return {'inserted': inserted, 'updated': updated, 'skipped': skipped}


def import_backup_workbook(data: bytes) -> Dict[str, Dict[str, int]]:
    sheets = parse_backup_workbook(data)
    summary: Dict[str, Dict[str, int]] = {}
    for table in IMPORT_ORDER:
        rows = sheets.get(table, [])
        if not rows:
            summary[table] = {'inserted': 0, 'updated': 0, 'skipped': 0}
            continue
        summary[table] = _upsert_rows(table, SHEET_SPECS[table], rows)
    db.conn.commit()
    return summary
